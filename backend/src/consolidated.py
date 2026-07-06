"""Сводная таблица по всем организациям (дневная / недельная).

ConsolidatedReporter.daily(day)      — строка = организация, столбцы = ФАКТ-показатели.
ConsolidatedReporter.weekly(any_date) — то же за ISO-неделю (Пн–Вс).

- Организации (6 шт.) — порядок из REPORT_TO_DIVISION_DESCRIPTION.
- Разные наборы показателей — бизнес-реальность, НЕ баг:
  Шаркер/ЦКР часто 0 (длинные ремонты); Мойки — без себестоимости/наценки.
- None / неприменимо → «—», без падений.
- Печатает выравненную таблицу и сохраняет xlsx в output/.
- Только чтение из 1С (через MetricsService), запись — только в xlsx.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Optional

from openpyxl import Workbook

try:
    from .odata_client import ODataClient
    from .references import REPORT_TO_DIVISION_DESCRIPTION
    from .metrics import MetricsService, DailyMetrics, WeeklyMetrics
except ImportError:  # запуск как скрипта
    from odata_client import ODataClient
    from references import REPORT_TO_DIVISION_DESCRIPTION
    from metrics import MetricsService, DailyMetrics, WeeklyMetrics


OUTPUT_DIR = Path(__file__).resolve().parent.parent / "output"
DASH = "—"

# Порядок организаций в сводной (report_name).
ORG_ORDER: list[str] = list(REPORT_TO_DIVISION_DESCRIPTION.keys())

# Столбцы сводной: (заголовок, ключ извлечения из строки-метрик).
COLUMNS: list[tuple[str, str]] = [
    ("Организация", "org"),
    ("Закрыто ЗН", "closed"),
    ("Нормочасы", "normhours"),
    ("Выработка/мастер", "output_master"),
    ("Выручка", "revenue"),
    ("Себест. ЗЧ", "cost"),
    ("Маржа", "margin"),
    ("Маржинальность %", "margin_pct"),
    ("Средний чек", "avg_check"),
    ("Наценка ЗЧ %", "markup_pct"),
    ("Незакр. ЗН >1дн", "unclosed"),
]


@dataclass
class ConsolidatedRow:
    """Строка сводной по одной организации. Значения — float | None."""
    org: str
    values: dict[str, Optional[float]]


class ConsolidatedReporter:
    def __init__(self, service: Optional[MetricsService] = None,
                 client: Optional[ODataClient] = None) -> None:
        self.service = service or MetricsService(client)

    # --------------------------------------------------------------- построение строк
    @staticmethod
    def _row_from_metrics(org: str, m) -> ConsolidatedRow:
        """Строит ConsolidatedRow из DailyMetrics или WeeklyMetrics.

        margin       = выручка − себест. (по факту);
        margin_pct   = маржа / выручка × 100 (при выручке > 0);
        avg_check    = выручка / закрыто ЗН (при закрытых > 0);
        None → прочерк (например, наценка при cost=0).
        """
        revenue = m.revenue.fact
        cost = m.cost.fact
        closed = m.closed_orders.fact

        margin = m.markup.fact  # revenue - cost (по факту)
        margin_pct = (margin / revenue * 100.0) if revenue else None
        avg_check = (revenue / closed) if closed else None

        values: dict[str, Optional[float]] = {
            "closed": closed,
            "normhours": m.normhours.fact,
            "output_master": m.output_per_master.fact,
            "revenue": revenue,
            "cost": cost if cost else None,
            "margin": margin,
            "margin_pct": margin_pct,
            "avg_check": avg_check,
            "markup_pct": m.markup_pct.fact,  # None при cost=0 (parts-only)
            "unclosed": m.unclosed_over_1day.fact,
        }
        return ConsolidatedRow(org=org, values=values)

    def daily(self, day: date) -> list[ConsolidatedRow]:
        """Сводная за день: печать + xlsx в output/."""
        rows: list[ConsolidatedRow] = []
        for org in ORG_ORDER:
            m = self.service.daily(org, day)
            rows.append(self._row_from_metrics(org, m))
        title = f"Сводная (дневная) за {day.strftime('%d.%m.%Y')}"
        self._print_table(title, rows)
        out = OUTPUT_DIR / f"Сводная_дневная_{day.isoformat()}.xlsx"
        self._save_xlsx(title, rows, out)
        print(f"\nСохранено: {out}")
        return rows

    def weekly(self, any_date: date) -> list[ConsolidatedRow]:
        """Сводная за ISO-неделю (Пн–Вс): печать + xlsx в output/."""
        rows: list[ConsolidatedRow] = []
        week_start = week_end = None
        for org in ORG_ORDER:
            m = self.service.weekly(org, any_date)
            week_start, week_end = m.week_start, m.week_end
            rows.append(self._row_from_metrics(org, m))
        label = f"{week_start.strftime('%d.%m')}–{week_end.strftime('%d.%m.%Y')}"
        title = f"Сводная (недельная) {label}"
        self._print_table(title, rows)
        tag = week_start.isoformat()
        out = OUTPUT_DIR / f"Сводная_недельная_{tag}.xlsx"
        self._save_xlsx(title, rows, out)
        print(f"\nСохранено: {out}")
        return rows

    # --------------------------------------------------------------- вывод
    @staticmethod
    def _fmt(key: str, value) -> str:
        if value is None:
            return DASH
        if key in ("closed", "unclosed"):
            return f"{value:,.0f}".replace(",", " ")
        if key in ("margin_pct", "markup_pct"):
            return f"{value:,.1f}".replace(",", " ")
        return f"{value:,.2f}".replace(",", " ")

    def _print_table(self, title: str, rows: list[ConsolidatedRow]) -> None:
        headers = [h for h, _ in COLUMNS]
        keys = [k for _, k in COLUMNS]

        # матрица строк (текст)
        table: list[list[str]] = []
        for r in rows:
            cells = [r.org]
            for k in keys[1:]:
                cells.append(self._fmt(k, r.values.get(k)))
            table.append(cells)

        widths = [len(h) for h in headers]
        for cells in table:
            for i, c in enumerate(cells):
                widths[i] = max(widths[i], len(c))

        print(f"\n{title}")
        line = "  ".join(h.ljust(widths[i]) if i == 0 else h.rjust(widths[i])
                          for i, h in enumerate(headers))
        print(line)
        print("-" * len(line))
        for cells in table:
            row_line = "  ".join(
                c.ljust(widths[i]) if i == 0 else c.rjust(widths[i])
                for i, c in enumerate(cells)
            )
            print(row_line)

    # --------------------------------------------------------------- xlsx
    def _save_xlsx(self, title: str, rows: list[ConsolidatedRow], out_path: Path) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "Сводная"

        headers = [h for h, _ in COLUMNS]
        keys = [k for _, k in COLUMNS]

        ws.cell(row=1, column=1, value=title)
        header_row = 3
        for ci, h in enumerate(headers, start=1):
            ws.cell(row=header_row, column=ci, value=h)

        for ri, r in enumerate(rows, start=header_row + 1):
            ws.cell(row=ri, column=1, value=r.org)
            for ci, k in enumerate(keys[1:], start=2):
                v = r.values.get(k)
                # числа — числовым типом; None → «—»
                ws.cell(row=ri, column=ci, value=(v if v is not None else DASH))

        out = Path(out_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out)
        return out
