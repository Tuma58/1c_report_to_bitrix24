"""ExcelReporter — заполнение шаблона отчёта дневными показателями.

Грузит КОПИЮ шаблона (backend/templates/report_template.xlsx), находит лист
по названию отчёта, пишет значения по карте ячеек из INCREMENT3.md и сохраняет
в out_path. Формульные ячейки (D16/F16, D17/F17, D18/F18, вся колонка H) НЕ
трогаются — их считает Excel.

Числа записываются числовым типом (float/int), дата — строкой ДД.ММ.ГГГГ.
Прочерк «—» — для несчитаемых значений.
"""
from __future__ import annotations

import os
import tempfile
from datetime import date
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

try:
    from .metrics import DailyMetrics, Metric
except ImportError:  # запуск как скрипта
    from metrics import DailyMetrics, Metric


# Путь к КОПИИ шаблона внутри проекта (self-contained).
TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "templates" / "report_template.xlsx"

# Название отчёта -> имя листа дневного отчёта в шаблоне.
REPORT_TO_SHEET: dict[str, str] = {
    "Арсенал": "Арсенал (Д)",
    "Реф. Сервис": "Реф.сервис (Д)",
}

# Название отчёта -> имя листа НЕДЕЛЬНОГО отчёта (Н) в шаблоне.
REPORT_TO_WEEKLY_SHEET: dict[str, str] = {
    "Арсенал": "Арсенал (Н)",
    "Реф. Сервис": "Реф.сервис (Н)",
}

DASH = "—"
REPORT_MANAGERS: dict[str, str] = {
    "Арсенал": "Мигунов А.А.",
    "Реф. Сервис": "Михеев А.Н.",
    "Шаркер": "Михеев А.Н.",
    "ЦКР": "Игонин А.П.",
    "Мойки": "Расторгуев С.В.",
    "Мойка на ул. Строителей": "Расторгуев С.В.",
    "Мойка на ул. Ульяновская": "Расторгуев С.В.",
}


class ExcelReporter:
    def __init__(self, template_path: Optional[Path] = None) -> None:
        self.template_path = Path(template_path) if template_path else TEMPLATE_PATH
        if not self.template_path.exists():
            raise FileNotFoundError(f"Шаблон не найден: {self.template_path}")

    @staticmethod
    def _num(value) -> Optional[float]:
        """Число как есть или None (для прочерка). None -> оставит DASH."""
        if value is None:
            return None
        return float(value)

    def _write(self, ws, coord: str, value) -> None:
        """Пишет число (числовым типом) либо прочерк «—» при None."""
        ws[coord] = value if value is not None else DASH

    @staticmethod
    def _pct(value) -> Optional[float]:
        """Преобразует процент из 1С/метрик (52) в Excel-долю (0.52)."""
        if value is None:
            return None
        return float(value) / 100.0

    @staticmethod
    def _diff_pct(current, previous) -> Optional[float]:
        """Процентное изменение: (текущее - предыдущее) / предыдущее * 100."""
        if current is None or previous in (None, 0):
            return None
        return (float(current) - float(previous)) / float(previous) * 100.0

    def _write_manager(self, ws, report_name: str) -> None:
        ws["D5"] = REPORT_MANAGERS.get(report_name, "")

    def fill_daily(
        self,
        report_name: str,
        day: date,
        metrics: DailyMetrics,
        out_path,
    ) -> Path:
        """Заполняет дневной лист шаблона и сохраняет в out_path.

        Не перезаписывает формульные ячейки (D16/F16, D17/F17, D18/F18, H*).
        """
        sheet_name = REPORT_TO_SHEET.get(report_name)
        if not sheet_name:
            raise ValueError(f"Нет маппинга листа для отчёта '{report_name}'")

        wb = load_workbook(self.template_path, keep_vba=False, data_only=False)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Лист '{sheet_name}' не найден в шаблоне")
        ws = wb[sheet_name]

        # Шапка: D3/D4 — широкая объединённая область D:I.
        ws["D3"] = report_name
        ws["D4"] = day.strftime("%d.%m.%Y")
        self._write_manager(ws, report_name)

        # Блок «Выработка» (D=ПЛАН / F=ФАКТ).
        self._write(ws, "D8", self._num(metrics.normhours.plan))
        self._write(ws, "F8", self._num(metrics.normhours.fact))
        self._write(ws, "D9", self._num(metrics.output_per_master.plan))
        self._write(ws, "F9", self._num(metrics.output_per_master.fact))
        self._write(ws, "D10", self._num(metrics.closed_orders.plan))
        self._write(ws, "F10", self._num(metrics.closed_orders.fact))

        # Блок «Финансы».
        self._write(ws, "D14", self._num(metrics.revenue.plan))
        self._write(ws, "F14", self._num(metrics.revenue.fact))
        self._write(ws, "D15", self._num(metrics.cost.plan))
        self._write(ws, "F15", self._num(metrics.cost.fact))
        # D16/F16 (Маржа), D17/F17 (Маржинальность), D18/F18 (Средний чек) —
        # ФОРМУЛЫ Excel, НЕ трогаем.
        self._write(ws, "D19", self._pct(metrics.markup_pct.plan))
        self._write(ws, "F19", self._pct(metrics.markup_pct.fact))

        # Блок «Незакрытые ЗН». D23 — левая верхняя ячейка объединённого D23:F23.
        self._write(ws, "D23", self._num(metrics.unclosed_over_1day.fact))
        # D24 — «ЗН без движения > 2 дней»: пока прочерк (нужен механизм снимков).
        ws["D24"] = DASH

        out = Path(out_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out)
        return out

    def fill_weekly(
        self,
        report_name: str,
        weekly_metrics,
        day_metrics: list,
        out_path,
    ) -> Path:
        """Заполняет НЕДЕЛЬНЫЙ лист (Н) шаблона и сохраняет в out_path.

        Разделы 1–2 (D=ПЛАН / F=ФАКТ) по карте ячеек INCREMENT5.md; раздел 3
        «Выручка по дням» — строки 24..30 (Пн..Вс) из day_metrics (7 шт.).

        НЕ трогает формульные ячейки: D17/F17 (Маржа), D18/F18 (Маржинальность),
        D19/F19 (Средний чек), I24:I31 (Ср.чек), D31/F31/H31 (ИТОГО), колонки H (%).
        C4 — неделя строкой «ДД.ММ–ДД.ММ.ГГГГ».
        """
        sheet_name = REPORT_TO_WEEKLY_SHEET.get(report_name)
        if not sheet_name:
            raise ValueError(f"Нет маппинга недельного листа для отчёта '{report_name}'")

        wb = load_workbook(self.template_path, keep_vba=False, data_only=False)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Лист '{sheet_name}' не найден в шаблоне")
        ws = wb[sheet_name]

        w = weekly_metrics

        # Шапка: D3/D4/D5 — широкая объединённая область D:I.
        ws["D3"] = report_name
        ws["D4"] = (
            f"{w.week_start.strftime('%d.%m')}–{w.week_end.strftime('%d.%m.%Y')}"
        )
        self._write_manager(ws, report_name)

        # Раздел 1 «Выработка».
        self._write(ws, "D9", self._num(w.normhours.plan))
        self._write(ws, "F9", self._num(w.normhours.fact))
        self._write(ws, "D10", self._num(w.output_per_master.plan))
        self._write(ws, "F10", self._num(w.output_per_master.fact))
        self._write(ws, "D11", self._num(w.closed_orders.plan))
        self._write(ws, "F11", self._num(w.closed_orders.fact))

        # Раздел 2 «Финансы».
        self._write(ws, "D15", self._num(w.revenue.plan))
        self._write(ws, "F15", self._num(w.revenue.fact))
        self._write(ws, "D16", self._num(w.cost.plan))
        self._write(ws, "F16", self._num(w.cost.fact))
        # D17/F17 (Маржа), D18/F18 (Маржинальность), D19/F19 (Средний чек) — ФОРМУЛЫ, НЕ трогаем.
        self._write(ws, "D20", self._pct(w.markup_pct.plan))
        self._write(ws, "F20", self._pct(w.markup_pct.fact))

        # Раздел 3 «Выручка по дням»: строка 24 = Пн (day_metrics[0]) .. строка 30 = Вс.
        # D=Выручка, F=Нормочасы, H=ЗН закрыто. I (Ср.чек), D31/F31/H31 (ИТОГО) — ФОРМУЛЫ.
        for i in range(7):
            row = 24 + i
            dm = day_metrics[i]
            self._write(ws, f"D{row}", self._num(dm.revenue.fact))
            self._write(ws, f"F{row}", self._num(dm.normhours.fact))
            self._write(ws, f"H{row}", self._num(dm.closed_orders.fact))

        out = Path(out_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out)
        return out


# ============================================================ ИНКРЕМЕНТ 6: МОЙКИ
WASH_DAILY_SHEET = "Мойка (Д)"
WASH_WEEKLY_SHEETS: dict[str, str] = {
    "Мойка на ул. Строителей": "Мойка ул. Строителей (Н)",
    "Мойка на ул. Ульяновская": "Мойка Ульяновская (Н)",
}


def fill_daily_wash(self, day, m_stroit, m_ulyan, out_path):
    """Заполняет лист «Мойка (Д)» (два блока) и сохраняет в out_path.

    Блок 1 «Строителей»: D8/F8 (машин), D9/F9 (выручка).
    Блок 2 «Ульяновская»: D14/F14 (машин), D15/F15 (выручка).
    НЕ трогает формулы Ср.чек (D10/F10, D16/F16) и H-колонки.
    """
    wb = load_workbook(self.template_path, keep_vba=False, data_only=False)
    if WASH_DAILY_SHEET not in wb.sheetnames:
        raise ValueError(f"Лист '{WASH_DAILY_SHEET}' не найден в шаблоне")
    ws = wb[WASH_DAILY_SHEET]

    # Шапка: D3/D4 — широкая объединённая область D:I.
    ws["D3"] = "Мойки"
    ws["D4"] = day.strftime("%d.%m.%Y")
    self._write_manager(ws, "Мойки")

    # Блок 1 — Мойка на ул. Строителей.
    self._write(ws, "D8", self._num(m_stroit.cars.plan))
    self._write(ws, "F8", self._num(m_stroit.cars.fact))
    self._write(ws, "D9", self._num(m_stroit.revenue.plan))
    self._write(ws, "F9", self._num(m_stroit.revenue.fact))

    # Блок 2 — Мойка на ул. Ульяновская.
    self._write(ws, "D14", self._num(m_ulyan.cars.plan))
    self._write(ws, "F14", self._num(m_ulyan.cars.fact))
    self._write(ws, "D15", self._num(m_ulyan.revenue.plan))
    self._write(ws, "F15", self._num(m_ulyan.revenue.fact))
    # D10/F10, D16/F16 (Ср.чек), H-колонки — ФОРМУЛЫ, НЕ трогаем.

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def fill_weekly_wash(self, report_name, weekly_metrics, day_breakdown, out_path):
    """Заполняет отдельный недельный лист мойки и сохраняет.

    Раздел 1: D9/F9 (машин недели), D10/F10 (выручка недели).
    Раздел 2 «Динамика по дням» строки 15..21 (Пн..Вс): D=машин, F=выручка,
    I=операторов (опц.). НЕ трогает формулы Ср.чек (D11/F11, H15..H21) и
    строку 22 ИТОГО (D22/F22/H22).
    """
    sheet_name = WASH_WEEKLY_SHEETS.get(report_name)
    if not sheet_name:
        raise ValueError(f"Нет недельного листа мойки для отчёта '{report_name}'")
    wb = load_workbook(self.template_path, keep_vba=False, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Лист '{sheet_name}' не найден в шаблоне")
    ws = wb[sheet_name]

    w = weekly_metrics

    ws["D3"] = report_name
    ws["D4"] = (
        f"{w.period_start.strftime('%d.%m')}–{w.period_end.strftime('%d.%m.%Y')}"
    )
    self._write_manager(ws, report_name)

    self._write(ws, "D9", self._num(w.cars.plan))
    self._write(ws, "F9", self._num(w.cars.fact))
    self._write(ws, "D10", self._num(w.revenue.plan))
    self._write(ws, "F10", self._num(w.revenue.fact))
    # D11/F11 (Ср.чек), H-колонки — ФОРМУЛЫ, НЕ трогаем.

    # Раздел 2 — динамика по дням: строка 15 = Пн (индекс 0) .. строка 21 = Вс.
    for i in range(7):
        row = 15 + i
        dm = day_breakdown[i]
        self._write(ws, f"D{row}", self._num(dm.cars.fact))
        self._write(ws, f"F{row}", self._num(dm.revenue.fact))
        # I — операторов за день (обе мойки); если нет — прочерк.
        operators = getattr(dm, "executors_count", 0)
        self._write(ws, f"I{row}", self._num(operators) if operators else None)
    # H15..H21 (Ср.чек), строка 22 ИТОГО (D22/F22/H22) — ФОРМУЛЫ, НЕ трогаем.

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


ExcelReporter.fill_daily_wash = fill_daily_wash
ExcelReporter.fill_weekly_wash = fill_weekly_wash


# ============================================================ ИНКРЕМЕНТ 7: ШАРКЕР/ЦКР
# Карты ячеек РАЗНЫЕ у Шаркер и ЦКР (у ЦКР сдвиг из-за строки 10 «Стоимость выработки»).
# Пайплайн пишется в колонку D (объединено D:F — пишем в D). Колонку H (формулы %) НЕ трогаем.
SHOP_DAILY_SHEET: dict[str, str] = {
    "Шаркер": "Шаркер (Д)",
    "ЦКР": "ЦКР (Д)",
}

# Для каждого отчёта — координаты по разделам.
SHOP_CELL_MAP: dict[str, dict] = {
    "Шаркер": {
        "normhours": ("D8", "F8"),      # план / факт из 1С
        "output": ("D9", "F9"),
        "output_cost": None,            # у Шаркера нет строки «Стоимость выработки»
        "active": "D13",
        "ready_today": "D14",
        "awaiting_parts": "D15",
        "overdue": "D16",
        "planned_close_week": "D17",
        "closed": ("D21", "F21"),
        "revenue": ("D22", "F22"),
        "payments_fact": None,
        "payments_plan": None,
        "insurance_count_fact": None,
        "insurance_sum_fact": None,
    },
    "ЦКР": {
        "normhours": ("D8", "F8"),
        "output": ("D9", "F9"),
        "output_cost": None,
        "active": "D14",
        "ready_today": "D15",
        "awaiting_parts": "D16",
        "overdue": "D17",
        "planned_close_week": "D18",
        "closed": ("D22", "F22"),
        "revenue": ("D23", "F23"),
        "payments_fact": None,
        "payments_plan": None,
        "insurance_count_fact": None,
        "insurance_sum_fact": None,
    },
}


def fill_daily_shop(self, report_name: str, shop_metrics, out_path):
    """Заполняет дневной лист длинного ремонта («Шаркер (Д)» или «ЦКР (Д)»).

    Пишет ПЛАН/ФАКТ выработки, пайплайн (колонка D), финансы. Показатели без
    подтверждённого источника (стоимость выработки, страховые ЗН) выводятся «—».
    Колонку H (формулы %) НЕ трогает.
    """
    sheet_name = SHOP_DAILY_SHEET.get(report_name)
    cmap = SHOP_CELL_MAP.get(report_name)
    if not sheet_name or not cmap:
        raise ValueError(f"Нет карты дневного листа для отчёта '{report_name}'")

    wb = load_workbook(self.template_path, keep_vba=False, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Лист '{sheet_name}' не найден в шаблоне")
    ws = wb[sheet_name]
    m = shop_metrics

    # Шапка: C3 — АТЦ, C4 — дата.
    ws["D3"] = report_name
    ws["D4"] = m.day.strftime("%d.%m.%Y")
    self._write_manager(ws, report_name)

    # ---- ВЫРАБОТКА (план / факт) ----
    d, f = cmap["normhours"]
    self._write(ws, d, self._num(m.normhours.plan))
    self._write(ws, f, self._num(m.normhours.fact))
    d, f = cmap["output"]
    self._write(ws, d, self._num(m.output_per_master.plan))
    self._write(ws, f, self._num(m.output_per_master.fact))
    if cmap.get("output_cost"):
        d, f = cmap["output_cost"]
        self._write(ws, d, self._num(m.output_cost.plan))
        self._write(ws, f, self._num(m.output_cost.fact))

    # ---- ПАЙПЛАЙН (колонка D, объединено D:F) ----
    self._write(ws, cmap["active"], self._num(m.active))
    self._write(ws, cmap["ready_today"], self._num(m.ready_today))
    self._write(ws, cmap["awaiting_parts"], self._num(m.awaiting_parts))
    self._write(ws, cmap["overdue"], self._num(m.overdue))
    self._write(ws, cmap["planned_close_week"], self._num(m.planned_close_week))

    # ---- ФИНАНСЫ ----
    d, f = cmap["closed"]
    self._write(ws, d, self._num(m.closed_orders.plan))
    self._write(ws, f, self._num(m.closed_orders.fact))
    d, f = cmap["revenue"]
    self._write(ws, d, self._num(m.revenue_closed.plan))
    self._write(ws, f, self._num(m.revenue_closed.fact))
    if cmap.get("payments_plan"):
        self._write(ws, cmap["payments_plan"], self._num(m.payments.plan))
    if cmap.get("payments_fact"):
        self._write(ws, cmap["payments_fact"], self._num(m.payments.fact))
    if cmap.get("insurance_count_fact"):
        self._write(ws, cmap["insurance_count_fact"], self._num(m.insurance_count.fact))
    if cmap.get("insurance_sum_fact"):
        self._write(ws, cmap["insurance_sum_fact"], self._num(m.insurance_sum.fact))
    # Колонка H (формулы %) — НЕ трогаем.

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


SHOP_WEEKLY_SHEET: dict[str, str] = {
    "Шаркер": "Шаркер (Н)",
    "ЦКР": "ЦКР (Н)",
}

SHOP_WEEKLY_CELL_MAP: dict[str, dict] = {
    "Шаркер": {
        "normhours": ("D9", "F9"),
        "output": ("D10", "F10"),
        "active": "D14",
        "closed": "D15",
        "revenue": "D16",
        "margin_pct": "D17",
        "avg_duration": "D18",
        "overdue": "D19",
        "awaiting_parts": "D20",
        "payments": "D21",
        "insurance_count": "D22",
        "insurance_sum": "D23",
        "daily_start": 27,
    },
    "ЦКР": {
        "normhours": ("D9", "F9"),
        "output": ("D10", "F10"),
        "active": "D14",
        "closed": "D15",
        "revenue": "D16",
        "margin_pct": "D17",
        "avg_duration": "D18",
        "overdue": "D19",
        "awaiting_parts": "D20",
        "payments": "D21",
        "insurance_count": None,
        "insurance_sum": None,
        "daily_start": 25,
    },
}


def _shop_masters(metric) -> Optional[float]:
    normhours = getattr(metric.normhours, "fact", None)
    output = getattr(metric.output_per_master, "fact", None)
    if normhours is None or not output:
        return None
    return normhours / output


def _shop_weekly_value(metric, key: str):
    if key == "active":
        return metric.active
    if key == "closed":
        return metric.closed_orders.fact
    if key == "revenue":
        return metric.revenue_closed.fact
    if key == "margin_pct":
        return metric.margin_pct.fact
    if key == "avg_duration":
        return metric.avg_duration_days.fact
    if key == "overdue":
        return metric.overdue
    if key == "awaiting_parts":
        return metric.awaiting_parts
    if key == "payments":
        return metric.payments.fact
    if key == "insurance_count":
        return metric.insurance_count.fact
    if key == "insurance_sum":
        return metric.insurance_sum.fact
    return None


def _day_label(metric, index: int) -> str:
    names = ("Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс")
    day = getattr(metric, "day", None)
    if day is None:
        return names[index]
    return f"{names[index]} {day.strftime('%d.%m.%Y')}"


def fill_weekly_shop(
    self,
    report_name: str,
    weekly_metrics,
    day_breakdown,
    out_path,
    prev_weekly_metrics=None,
):
    """Заполняет недельный лист длинного ремонта («Шаркер (Н)» или «ЦКР (Н)»)."""
    sheet_name = SHOP_WEEKLY_SHEET.get(report_name)
    cmap = SHOP_WEEKLY_CELL_MAP.get(report_name)
    if not sheet_name or not cmap:
        raise ValueError(f"Нет карты недельного листа для отчёта '{report_name}'")

    wb = load_workbook(self.template_path, keep_vba=False, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Лист '{sheet_name}' не найден в шаблоне")
    ws = wb[sheet_name]
    m = weekly_metrics

    ws["D3"] = report_name
    ws["D4"] = f"{m.week_start.strftime('%d.%m')}–{m.week_end.strftime('%d.%m.%Y')}"
    self._write_manager(ws, report_name)

    d, f = cmap["normhours"]
    self._write(ws, d, self._num(m.normhours.plan))
    self._write(ws, f, self._num(m.normhours.fact))
    d, f = cmap["output"]
    self._write(ws, d, self._num(m.output_per_master.plan))
    self._write(ws, f, self._num(m.output_per_master.fact))

    for key in (
        "active",
        "closed",
        "revenue",
        "margin_pct",
        "avg_duration",
        "overdue",
        "awaiting_parts",
        "payments",
        "insurance_count",
        "insurance_sum",
    ):
        cell = cmap.get(key)
        if not cell:
            continue
        current = _shop_weekly_value(m, key)
        previous = _shop_weekly_value(prev_weekly_metrics, key) if prev_weekly_metrics else None
        writer = self._pct if key == "margin_pct" else self._num
        self._write(ws, cell, writer(current))
        self._write(ws, f"H{cell[1:]}", writer(previous))
        self._write(ws, f"I{cell[1:]}", self._pct(self._diff_pct(current, previous)))

    start_row = cmap["daily_start"]
    for i in range(7):
        row = start_row + i
        dm = day_breakdown[i]
        ws[f"B{row}"] = _day_label(dm, i)
        self._write(ws, f"D{row}", self._num(dm.normhours.fact))
        self._write(ws, f"F{row}", self._num(_shop_masters(dm)))

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


ExcelReporter.fill_daily_shop = fill_daily_shop
ExcelReporter.fill_weekly_shop = fill_weekly_shop


# ============================================================ ЕДИНЫЙ ФАЙЛ
# Эти методы заполняют уже открытую книгу. Старые fill_* методы оставлены для
# smoke-тестов и точечной генерации, но основной CLI теперь сохраняет один xlsx
# с отдельными листами, как в исходном шаблоне.
def new_workbook(self):
    return load_workbook(self.template_path, keep_vba=False, data_only=False)


def save_workbook(self, wb, out_path) -> Path:
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(prefix=f".{out.name}.", suffix=".xlsx", dir=str(out.parent))
    os.close(fd)
    tmp = Path(tmp_name)
    try:
        wb.save(tmp)
        tmp.replace(out)
    finally:
        if tmp.exists():
            tmp.unlink()
    return out


def fill_daily_in_workbook(self, wb, report_name: str, day: date, metrics: DailyMetrics) -> None:
    sheet_name = REPORT_TO_SHEET.get(report_name)
    if not sheet_name or sheet_name not in wb.sheetnames:
        raise ValueError(f"Лист дневного отчёта '{report_name}' не найден")
    ws = wb[sheet_name]

    ws["D3"] = report_name
    ws["D4"] = day.strftime("%d.%m.%Y")
    self._write_manager(ws, report_name)
    self._write(ws, "D8", self._num(metrics.normhours.plan))
    self._write(ws, "F8", self._num(metrics.normhours.fact))
    self._write(ws, "D9", self._num(metrics.output_per_master.plan))
    self._write(ws, "F9", self._num(metrics.output_per_master.fact))
    self._write(ws, "D10", self._num(metrics.closed_orders.plan))
    self._write(ws, "F10", self._num(metrics.closed_orders.fact))
    self._write(ws, "D14", self._num(metrics.revenue.plan))
    self._write(ws, "F14", self._num(metrics.revenue.fact))
    self._write(ws, "D15", self._num(metrics.cost.plan))
    self._write(ws, "F15", self._num(metrics.cost.fact))
    self._write(ws, "D19", self._pct(metrics.markup_pct.plan))
    self._write(ws, "F19", self._pct(metrics.markup_pct.fact))
    self._write(ws, "D23", self._num(metrics.unclosed_over_1day.fact))
    ws["D24"] = DASH


def fill_weekly_in_workbook(self, wb, report_name: str, weekly_metrics, day_metrics: list) -> None:
    sheet_name = REPORT_TO_WEEKLY_SHEET.get(report_name)
    if not sheet_name or sheet_name not in wb.sheetnames:
        raise ValueError(f"Лист недельного отчёта '{report_name}' не найден")
    ws = wb[sheet_name]
    w = weekly_metrics

    ws["D3"] = report_name
    ws["D4"] = f"{w.week_start.strftime('%d.%m')}–{w.week_end.strftime('%d.%m.%Y')}"
    self._write_manager(ws, report_name)
    self._write(ws, "D9", self._num(w.normhours.plan))
    self._write(ws, "F9", self._num(w.normhours.fact))
    self._write(ws, "D10", self._num(w.output_per_master.plan))
    self._write(ws, "F10", self._num(w.output_per_master.fact))
    self._write(ws, "D11", self._num(w.closed_orders.plan))
    self._write(ws, "F11", self._num(w.closed_orders.fact))
    self._write(ws, "D15", self._num(w.revenue.plan))
    self._write(ws, "F15", self._num(w.revenue.fact))
    self._write(ws, "D16", self._num(w.cost.plan))
    self._write(ws, "F16", self._num(w.cost.fact))
    self._write(ws, "D20", self._pct(w.markup_pct.plan))
    self._write(ws, "F20", self._pct(w.markup_pct.fact))

    for i in range(7):
        row = 24 + i
        dm = day_metrics[i]
        self._write(ws, f"D{row}", self._num(dm.revenue.fact))
        self._write(ws, f"F{row}", self._num(dm.normhours.fact))
        self._write(ws, f"H{row}", self._num(dm.closed_orders.fact))


def fill_daily_wash_in_workbook(self, wb, day, m_stroit, m_ulyan) -> None:
    if WASH_DAILY_SHEET not in wb.sheetnames:
        raise ValueError(f"Лист '{WASH_DAILY_SHEET}' не найден")
    ws = wb[WASH_DAILY_SHEET]

    ws["D3"] = "Мойки"
    ws["D4"] = day.strftime("%d.%m.%Y")
    self._write_manager(ws, "Мойки")
    self._write(ws, "D8", self._num(m_stroit.cars.plan))
    self._write(ws, "F8", self._num(m_stroit.cars.fact))
    self._write(ws, "D9", self._num(m_stroit.revenue.plan))
    self._write(ws, "F9", self._num(m_stroit.revenue.fact))
    self._write(ws, "D14", self._num(m_ulyan.cars.plan))
    self._write(ws, "F14", self._num(m_ulyan.cars.fact))
    self._write(ws, "D15", self._num(m_ulyan.revenue.plan))
    self._write(ws, "F15", self._num(m_ulyan.revenue.fact))


def fill_weekly_wash_in_workbook(self, wb, report_name, weekly_metrics, day_breakdown) -> None:
    sheet_name = WASH_WEEKLY_SHEETS.get(report_name)
    if not sheet_name:
        raise ValueError(f"Нет недельного листа мойки для отчёта '{report_name}'")
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Лист '{sheet_name}' не найден")
    ws = wb[sheet_name]
    w = weekly_metrics

    ws["D3"] = report_name
    ws["D4"] = f"{w.period_start.strftime('%d.%m')}–{w.period_end.strftime('%d.%m.%Y')}"
    self._write_manager(ws, report_name)
    self._write(ws, "D9", self._num(w.cars.plan))
    self._write(ws, "F9", self._num(w.cars.fact))
    self._write(ws, "D10", self._num(w.revenue.plan))
    self._write(ws, "F10", self._num(w.revenue.fact))
    for i in range(7):
        row = 15 + i
        dm = day_breakdown[i]
        self._write(ws, f"D{row}", self._num(dm.cars.fact))
        self._write(ws, f"F{row}", self._num(dm.revenue.fact))
        operators = getattr(dm, "executors_count", 0)
        self._write(ws, f"I{row}", self._num(operators) if operators else None)


def fill_daily_shop_in_workbook(self, wb, report_name: str, shop_metrics) -> None:
    sheet_name = SHOP_DAILY_SHEET.get(report_name)
    cmap = SHOP_CELL_MAP.get(report_name)
    if not sheet_name or sheet_name not in wb.sheetnames or not cmap:
        raise ValueError(f"Лист дневного отчёта '{report_name}' не найден")
    ws = wb[sheet_name]
    m = shop_metrics

    ws["D3"] = report_name
    ws["D4"] = m.day.strftime("%d.%m.%Y")
    self._write_manager(ws, report_name)
    d, f = cmap["normhours"]
    self._write(ws, d, self._num(m.normhours.plan))
    self._write(ws, f, self._num(m.normhours.fact))
    d, f = cmap["output"]
    self._write(ws, d, self._num(m.output_per_master.plan))
    self._write(ws, f, self._num(m.output_per_master.fact))
    if cmap.get("output_cost"):
        d, f = cmap["output_cost"]
        self._write(ws, d, self._num(m.output_cost.plan))
        self._write(ws, f, self._num(m.output_cost.fact))
    self._write(ws, cmap["active"], self._num(m.active))
    self._write(ws, cmap["ready_today"], self._num(m.ready_today))
    self._write(ws, cmap["awaiting_parts"], self._num(m.awaiting_parts))
    self._write(ws, cmap["overdue"], self._num(m.overdue))
    self._write(ws, cmap["planned_close_week"], self._num(m.planned_close_week))
    d, f = cmap["closed"]
    self._write(ws, d, self._num(m.closed_orders.plan))
    self._write(ws, f, self._num(m.closed_orders.fact))
    d, f = cmap["revenue"]
    self._write(ws, d, self._num(m.revenue_closed.plan))
    self._write(ws, f, self._num(m.revenue_closed.fact))
    if cmap.get("payments_plan"):
        self._write(ws, cmap["payments_plan"], self._num(m.payments.plan))
    if cmap.get("payments_fact"):
        self._write(ws, cmap["payments_fact"], self._num(m.payments.fact))
    if cmap.get("insurance_count_fact"):
        self._write(ws, cmap["insurance_count_fact"], self._num(m.insurance_count.fact))
    if cmap.get("insurance_sum_fact"):
        self._write(ws, cmap["insurance_sum_fact"], self._num(m.insurance_sum.fact))


def fill_weekly_shop_in_workbook(
    self,
    wb,
    report_name: str,
    weekly_metrics,
    day_breakdown,
    prev_weekly_metrics=None,
) -> None:
    sheet_name = SHOP_WEEKLY_SHEET.get(report_name)
    cmap = SHOP_WEEKLY_CELL_MAP.get(report_name)
    if not sheet_name or sheet_name not in wb.sheetnames or not cmap:
        raise ValueError(f"Лист недельного отчёта '{report_name}' не найден")
    ws = wb[sheet_name]
    m = weekly_metrics

    ws["D3"] = report_name
    ws["D4"] = f"{m.week_start.strftime('%d.%m')}–{m.week_end.strftime('%d.%m.%Y')}"
    self._write_manager(ws, report_name)
    d, f = cmap["normhours"]
    self._write(ws, d, self._num(m.normhours.plan))
    self._write(ws, f, self._num(m.normhours.fact))
    d, f = cmap["output"]
    self._write(ws, d, self._num(m.output_per_master.plan))
    self._write(ws, f, self._num(m.output_per_master.fact))
    for key in (
        "active",
        "closed",
        "revenue",
        "margin_pct",
        "avg_duration",
        "overdue",
        "awaiting_parts",
        "payments",
        "insurance_count",
        "insurance_sum",
    ):
        cell = cmap.get(key)
        if not cell:
            continue
        current = _shop_weekly_value(m, key)
        previous = _shop_weekly_value(prev_weekly_metrics, key) if prev_weekly_metrics else None
        writer = self._pct if key == "margin_pct" else self._num
        self._write(ws, cell, writer(current))
        self._write(ws, f"H{cell[1:]}", writer(previous))
        self._write(ws, f"I{cell[1:]}", self._pct(self._diff_pct(current, previous)))
    for i in range(7):
        row = cmap["daily_start"] + i
        dm = day_breakdown[i]
        ws[f"B{row}"] = _day_label(dm, i)
        self._write(ws, f"D{row}", self._num(dm.normhours.fact))
        self._write(ws, f"F{row}", self._num(_shop_masters(dm)))


ExcelReporter.new_workbook = new_workbook
ExcelReporter.save_workbook = save_workbook
ExcelReporter.fill_daily_in_workbook = fill_daily_in_workbook
ExcelReporter.fill_weekly_in_workbook = fill_weekly_in_workbook
ExcelReporter.fill_daily_wash_in_workbook = fill_daily_wash_in_workbook
ExcelReporter.fill_weekly_wash_in_workbook = fill_weekly_wash_in_workbook
ExcelReporter.fill_daily_shop_in_workbook = fill_daily_shop_in_workbook
ExcelReporter.fill_weekly_shop_in_workbook = fill_weekly_shop_in_workbook
