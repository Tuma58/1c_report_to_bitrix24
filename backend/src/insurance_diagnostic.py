"""Диагностическая книга по неврученным страховым ЗН."""
from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from .excel_reporter import ExcelReporter
    from .generate_reports import OUTPUT_DIR, generate_workbook
    from .metrics import MetricsService
    from .odata_client import ODataClient
    from .references import REPORT_TO_DIVISION_DESCRIPTION, References
    from .repositories import InsuranceRepository, _dt
except ImportError:
    from excel_reporter import ExcelReporter
    from generate_reports import OUTPUT_DIR, generate_workbook
    from metrics import MetricsService
    from odata_client import ODataClient
    from references import REPORT_TO_DIVISION_DESCRIPTION, References
    from repositories import InsuranceRepository, _dt


REPORTS = ("Шаркер", "ЦКР")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")


@dataclass
class SelectionResult:
    report_name: str
    rows: list[dict]
    counters: dict[str, int | float | str]


def _parse_date(raw: str) -> date:
    return datetime.strptime(raw, "%Y-%m-%d").date()


def _invoice_text(invoices: list[dict], delivered_keys: set[str]) -> str:
    values = []
    for inv in invoices:
        number = inv.get("Number") or ""
        date_raw = inv.get("Date") or ""
        mark = "дата вручения есть" if inv.get("Ref_Key") in delivered_keys else "даты вручения нет"
        values.append(f"{number} от {date_raw} ({mark})")
    return "\n".join(values)


def _select_orders(
    insurance: InsuranceRepository,
    refs: References,
    report_name: str,
    as_of: date,
) -> SelectionResult:
    division_key = refs.division_key(report_name)
    if not division_key:
        raise RuntimeError(f"Не найдено подразделение для {report_name}")

    orders = insurance._closed_insurance_orders(division_key, as_of)
    invoices_by_order = insurance._invoices_by_order(division_key, as_of)
    delivered_invoice_keys = insurance._delivered_invoice_keys(as_of)
    balance_by_order = insurance._balance_by_order(as_of)

    rows: list[dict] = []
    counters: dict[str, int | float | str] = {
        "report_name": report_name,
        "division_description": REPORT_TO_DIVISION_DESCRIPTION.get(report_name, ""),
        "division_key": division_key,
        "candidate_orders": len(orders),
        "balance_order_links": len(balance_by_order),
        "invoice_order_links": len(invoices_by_order),
        "delivered_invoice_links": len(delivered_invoice_keys),
        "no_positive_balance_excluded": 0,
        "delivered_excluded": 0,
        "selected_without_invoice": 0,
        "selected_without_delivery_date": 0,
        "selected_count": 0,
        "selected_sum": 0.0,
    }

    for order in orders:
        order_key = order.get("Ref_Key")
        invoices = invoices_by_order.get(order_key, [])
        balance = balance_by_order.get(order_key, 0.0)
        if balance <= insurance.PAYMENT_TOLERANCE:
            counters["no_positive_balance_excluded"] = int(counters["no_positive_balance_excluded"]) + 1
            continue

        delivered = any(inv.get("Ref_Key") in delivered_invoice_keys for inv in invoices)
        if delivered:
            counters["delivered_excluded"] = int(counters["delivered_excluded"]) + 1
            continue

        reason = "Нет счет-фактуры"
        if invoices:
            reason = "Счет-фактура есть, дата вручения КА не заполнена"
            counters["selected_without_delivery_date"] = int(counters["selected_without_delivery_date"]) + 1
        else:
            counters["selected_without_invoice"] = int(counters["selected_without_invoice"]) + 1

        counters["selected_count"] = int(counters["selected_count"]) + 1
        counters["selected_sum"] = float(counters["selected_sum"]) + balance
        rows.append(
            {
                "Направление": report_name,
                "Номер ЗН": order.get("Number"),
                "Ref_Key ЗН": order_key,
                "Дата закрытия": order.get("ДатаЗакрытия"),
                "Сумма ЗН": order.get("СуммаДокумента"),
                "Неоплаченный остаток": balance,
                "Причина отбора": reason,
                "Количество СФ": len(invoices),
                "Счет-фактуры": _invoice_text(invoices, delivered_invoice_keys),
                "Вид ремонта Key": order.get("ВидРемонта_Key"),
            }
        )

    return SelectionResult(report_name, rows, counters)


def _append_table(ws, headers: list[str], rows: list[list[object]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col_idx, _header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=row_idx, column=col_idx).value or ""))
            for row_idx in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 55)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _add_selection_sheet(wb, result: SelectionResult) -> None:
    title = f"ЗН {result.report_name}"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    headers = [
        "Направление",
        "Номер ЗН",
        "Ref_Key ЗН",
        "Дата закрытия",
        "Сумма ЗН",
        "Неоплаченный остаток",
        "Причина отбора",
        "Количество СФ",
        "Счет-фактуры",
        "Вид ремонта Key",
    ]
    rows = [[row.get(header, "") for header in headers] for row in result.rows]
    _append_table(ws, headers, rows)
    for row in ws.iter_rows(min_row=2):
        if row[6].value == "Нет счет-фактуры":
            row[6].fill = WARN_FILL


def _add_algorithm_sheet(wb, results: list[SelectionResult], insurance: InsuranceRepository, as_of: date) -> None:
    title = "Алгоритм страховые ЗН"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    rows: list[list[object]] = [
        ["Дата отчета", as_of.isoformat()],
        ["Время формирования", insurance._selection_end(as_of).isoformat()],
        ["Нижняя граница закрытия ЗН", insurance.SELECTION_START.isoformat()],
        ["Верхняя граница закрытия ЗН", f"{insurance._selection_end(as_of).isoformat()} не включительно"],
        ["Виды ремонта", "; ".join(insurance.REPAIR_TYPES)],
        ["Порог неоплаченного остатка", insurance.PAYMENT_TOLERANCE],
        ["Шаг 1", "Берем закрытые ЗН выбранного подразделения с нужными видами ремонта."],
        ["Шаг 2", "Оставляем только ЗН с положительным конечным остатком взаиморасчетов СуммаBalance > 1."],
        ["Шаг 3", "Ищем подчиненные счет-фактуры по ДокументОснование = Ref_Key ЗН."],
        ["Шаг 4", "Если счет-фактуры нет, ЗН попадает в выборку."],
        ["Шаг 5", "Если счет-фактура есть и по ней заполнена Дата2/Получен КА, ЗН исключается."],
        ["Шаг 6", "Если счет-фактура есть, но дата вручения КА не заполнена, ЗН попадает в выборку."],
        ["Итог количества", "Количество строк на листах ЗН Шаркер и ЗН ЦКР."],
        ["Итог суммы", "Сумма поля Неоплаченный остаток / СуммаBalance по найденным ЗН."],
    ]
    for result in results:
        rows.extend(
            [
                ["", ""],
                [f"{result.report_name}: кандидатов", result.counters["candidate_orders"]],
                [f"{result.report_name}: найдено", result.counters["selected_count"]],
                [f"{result.report_name}: сумма", result.counters["selected_sum"]],
                [f"{result.report_name}: исключено без положительного остатка", result.counters["no_positive_balance_excluded"]],
                [f"{result.report_name}: исключено с датой вручения КА", result.counters["delivered_excluded"]],
                [f"{result.report_name}: выбрано без СФ", result.counters["selected_without_invoice"]],
                [f"{result.report_name}: выбрано с СФ без даты", result.counters["selected_without_delivery_date"]],
            ]
        )
    _append_table(ws, ["Параметр", "Значение"], rows)


def _add_filters_sheet(
    wb,
    results: list[SelectionResult],
    insurance: InsuranceRepository,
    refs: References,
    as_of: date,
) -> None:
    title = "Фильтры страховые ЗН"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    rows: list[list[object]] = []
    closed_key = refs.statuses().get("Закрыт", "")
    repair_clause = " or ".join(f"ВидРемонта_Key eq guid'{key}'" for key in sorted(insurance._repair_type_keys()))
    for result in results:
        division_key = result.counters["division_key"]
        order_filter = (
            f"Состояние_Key eq guid'{closed_key}' "
            f"and ПодразделениеКомпании_Key eq guid'{division_key}' "
            f"and ДатаЗакрытия ge {_dt(insurance.SELECTION_START)} "
            f"and ДатаЗакрытия lt {_dt(insurance._selection_end(as_of))} "
            f"and ({repair_clause})"
        )
        invoice_filter = (
            f"ПодразделениеКомпании_Key eq guid'{division_key}' "
            f"and Date lt {_dt(insurance._selection_end(as_of))} "
            f"and ДокументОснование_Type eq '{insurance.INVOICE_ORDER_TYPE}'"
        )
        rows.extend(
            [
                [result.report_name, "Document_ЗаказНаряд", order_filter, "Ref_Key,Number,ДатаЗакрытия,ВидРемонта_Key,СуммаДокумента"],
                [
                    result.report_name,
                    insurance.INVOICE_DOC,
                    invoice_filter,
                    "Ref_Key,Number,Date,ДокументОснование,ДокументОснование_Type,СуммаДокумента,Выставлен",
                ],
            ]
        )
    rows.extend(
        [
            [
                "Общий",
                insurance._balance_entity(as_of),
                f"Сделка_Type eq '{insurance.PAYMENT_ORDER_TYPE}' and СуммаBalance gt {insurance.PAYMENT_TOLERANCE:g}",
                "Контрагент_Key,ДоговорВзаиморасчетов_Key,Сделка,Сделка_Type,СуммаBalance",
            ],
            [
                "Общий",
                insurance.ORIGINALS_REGISTER,
                (
                    f"ДокументСсылка_Type eq '{insurance.INVOICE_ORIGINAL_TYPE}' "
                    "and ОригиналПолучен eq true "
                    "and Дата2 gt datetime'1901-01-01T00:00:00' "
                    f"and Дата2 lt {_dt(insurance._selection_end(as_of))}"
                ),
                f"ДокументСсылка,ДокументСсылка_Type,ОригиналПолучен,Дата,{insurance.DELIVERY_DATE_FIELD}",
            ],
        ]
    )
    _append_table(ws, ["Направление", "OData сущность", "$filter", "$select"], rows)


def build_diagnostic_workbook(as_of: date) -> Path:
    generated_at = datetime.now().replace(microsecond=0)
    service = MetricsService()
    service.insurance.selection_end_at = generated_at
    reporter = ExcelReporter()
    generated = generate_workbook(service, reporter, "all", as_of, as_of)
    out_path = generated[0].path

    client = ODataClient()
    refs = References(client)
    insurance = InsuranceRepository(client, refs)
    insurance.selection_end_at = generated_at
    results = [_select_orders(insurance, refs, report_name, as_of) for report_name in REPORTS]

    wb = load_workbook(out_path)
    for result in results:
        _add_selection_sheet(wb, result)
    _add_algorithm_sheet(wb, results, insurance, as_of)
    _add_filters_sheet(wb, results, insurance, refs, as_of)

    stamp = generated_at.strftime("%Y-%m-%d_%H-%M-%S")
    diagnostic_path = OUTPUT_DIR / f"reports_with_insurance_diagnostic_{stamp}.xlsx"
    wb.save(diagnostic_path)
    return diagnostic_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Создать тестовую книгу с диагностикой страховых ЗН")
    parser.add_argument("--date", type=_parse_date, default=date.today())
    args = parser.parse_args()
    out = build_diagnostic_workbook(args.date)
    print(out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
