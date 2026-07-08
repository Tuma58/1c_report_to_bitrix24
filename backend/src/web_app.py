"""Минимальный web GUI для настройки и запуска отчётов.

Работает без внешнего web-фреймворка: стандартный ThreadingHTTPServer.
По умолчанию слушает 127.0.0.1:8080. Для VPS можно задать WEB_UI_HOST,
WEB_UI_PORT и WEB_UI_PASSWORD в backend/.env.
"""
from __future__ import annotations

import base64
import argparse
import hashlib
import html
import json
import os
import re
import secrets
import sys
import tempfile
import time
from datetime import date, datetime, timedelta
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Optional
from urllib.parse import parse_qs, quote, urlencode, unquote, urlparse

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    from .bitrix_sender import BitrixError
    from .email_sender import EmailError
    from .generate_reports import (
        OUTPUT_DIR,
        GeneratedReport,
        generate_workbook,
        send_to_email,
        send_to_bitrix,
    )
    from .excel_reporter import ExcelReporter
    from .metrics import MetricsService
    from .odata_client import ODataError, ODataUnavailableError
    from .scheduler import ScheduleError, install_from_values as install_schedule
except ImportError:  # запуск как скрипта из backend/src
    from bitrix_sender import BitrixError
    from email_sender import EmailError
    from generate_reports import (
        OUTPUT_DIR,
        GeneratedReport,
        generate_workbook,
        send_to_email,
        send_to_bitrix,
    )
    from excel_reporter import ExcelReporter
    from metrics import MetricsService
    from odata_client import ODataError, ODataUnavailableError
    from scheduler import ScheduleError, install_from_values as install_schedule


APP_TITLE = "Отчёты АТЦ"
BACKEND_DIR = Path(__file__).resolve().parent.parent
ENV_PATH = BACKEND_DIR / ".env"
ENV_EXAMPLE_PATH = BACKEND_DIR / ".env.example"
USERS_PATH = BACKEND_DIR / "users.json"
PASSWORD_HASH_ITERATIONS = 260_000
MODES = {
    "all": "Все листы",
    "daily": "День",
    "weekly": "Неделя",
}
SETTINGS_FIELDS = [
    {
        "key": "ODATA_BASE_URL",
        "label": "URL 1С OData",
        "section": "1С OData",
        "help": "Например: http://host/base/odata/standard.odata/",
    },
    {
        "key": "ODATA_LOGIN",
        "label": "Логин 1С",
        "section": "1С OData",
        "autocomplete": "username",
    },
    {
        "key": "ODATA_PASSWORD",
        "label": "Пароль 1С",
        "section": "1С OData",
        "secret": True,
        "autocomplete": "current-password",
    },
    {
        "key": "ODATA_TIMEOUT",
        "label": "Таймаут OData, сек.",
        "section": "1С OData",
        "kind": "int",
        "min": 1,
        "max": 600,
    },
    {
        "key": "ODATA_RETRIES",
        "label": "Повторы OData",
        "section": "1С OData",
        "kind": "int",
        "min": 1,
        "max": 20,
    },
    {
        "key": "BITRIX_WEBHOOK_URL",
        "label": "Webhook Bitrix24",
        "section": "Bitrix24",
        "secret": True,
        "help": "URL вида https://portal.bitrix24.ru/rest/<user>/<token>/",
        "autocomplete": "off",
    },
    {
        "key": "BITRIX_CHAT_ID",
        "label": "ID чата Bitrix24",
        "section": "Bitrix24",
        "help": "Например: chat123",
    },
    {
        "key": "BITRIX_DISK_FOLDER_ID",
        "label": "ID папки Диска Bitrix24",
        "section": "Bitrix24",
    },
    {
        "key": "EMAIL_SMTP_HOST",
        "label": "SMTP сервер",
        "section": "Email",
        "help": "Например: smtp.yandex.ru или smtp.gmail.com",
    },
    {
        "key": "EMAIL_SMTP_PORT",
        "label": "SMTP порт",
        "section": "Email",
        "kind": "int",
        "min": 1,
        "max": 65535,
    },
    {
        "key": "EMAIL_SMTP_LOGIN",
        "label": "SMTP логин",
        "section": "Email",
        "autocomplete": "username",
    },
    {
        "key": "EMAIL_SMTP_PASSWORD",
        "label": "SMTP пароль",
        "section": "Email",
        "secret": True,
        "autocomplete": "new-password",
    },
    {
        "key": "EMAIL_FROM",
        "label": "Email отправителя",
        "section": "Email",
    },
    {
        "key": "EMAIL_TO",
        "label": "Получатели",
        "section": "Email",
        "help": "Несколько адресов можно разделить запятыми или точками с запятой.",
    },
    {
        "key": "EMAIL_USE_TLS",
        "label": "Использовать STARTTLS",
        "section": "Email",
        "kind": "bool",
    },
    {
        "key": "EMAIL_USE_SSL",
        "label": "Использовать SMTP SSL",
        "section": "Email",
        "kind": "bool",
    },
    {
        "key": "SCHEDULE_ENABLED",
        "label": "Включить автоматическую генерацию",
        "section": "Расписание",
        "kind": "bool",
    },
    {
        "key": "SCHEDULE_DAILY_ENABLED",
        "label": "Дневной отчёт",
        "section": "Расписание",
        "kind": "bool",
    },
    {
        "key": "SCHEDULE_DAILY_TIME",
        "label": "Время дневного отчёта",
        "section": "Расписание",
        "kind": "time",
    },
    {
        "key": "SCHEDULE_WEEKLY_ENABLED",
        "label": "Недельный отчёт",
        "section": "Расписание",
        "kind": "bool",
    },
    {
        "key": "SCHEDULE_WEEKLY_DAY",
        "label": "День недельного отчёта",
        "section": "Расписание",
        "kind": "select",
        "options": [
            ("1", "Понедельник"),
            ("2", "Вторник"),
            ("3", "Среда"),
            ("4", "Четверг"),
            ("5", "Пятница"),
            ("6", "Суббота"),
            ("0", "Воскресенье"),
        ],
    },
    {
        "key": "SCHEDULE_WEEKLY_TIME",
        "label": "Время недельного отчёта",
        "section": "Расписание",
        "kind": "time",
    },
    {
        "key": "SCHEDULE_SEND_BITRIX",
        "label": "Рассылать в Bitrix24",
        "section": "Расписание",
        "kind": "bool",
    },
    {
        "key": "SCHEDULE_SEND_EMAIL",
        "label": "Рассылать по email",
        "section": "Расписание",
        "kind": "bool",
    },
    {
        "key": "WEB_UI_HOST",
        "label": "Адрес Web UI",
        "section": "Web UI",
        "help": "0.0.0.0 для доступа извне, 127.0.0.1 для reverse proxy.",
    },
    {
        "key": "WEB_UI_PORT",
        "label": "Порт Web UI",
        "section": "Web UI",
        "kind": "port",
        "help": "Порты 80 и 443 запрещены для этого приложения.",
    },
    {
        "key": "WEB_UI_USER",
        "label": "Логин Web UI",
        "section": "Web UI",
        "autocomplete": "username",
    },
    {
        "key": "WEB_UI_PASSWORD",
        "label": "Пароль Web UI",
        "section": "Web UI",
        "secret": True,
        "autocomplete": "new-password",
    },
]
SETTINGS_KEYS = [field["key"] for field in SETTINGS_FIELDS]


def _env_int(name: str, default: int) -> int:
    raw = os.getenv(name, "").strip()
    if not raw:
        return default
    try:
        return int(raw)
    except ValueError:
        return default


def _parse_report_date(raw: str) -> date:
    return datetime.strptime(raw, "%Y-%m-%d").date()


def _default_date() -> date:
    return date.today() - timedelta(days=1)


def _file_url(path: Path) -> str:
    return f"/download?file={quote(path.name)}"


def _view_url(path: Path, sheet: str = "") -> str:
    query = {"file": path.name}
    if sheet:
        query["sheet"] = sheet
    return f"/view?{urlencode(query)}"


def _view_location(path: Path, *, sheet: str = "", message: str = "", error: str = "") -> str:
    query = {"file": path.name}
    if sheet:
        query["sheet"] = sheet
    if message:
        query["message"] = message
    if error:
        query["error"] = error
    return f"/view?{urlencode(query)}"


def _safe_output_path(filename: str) -> Optional[Path]:
    if not filename or "/" in filename or "\\" in filename or not filename.endswith(".xlsx"):
        return None
    path = (OUTPUT_DIR / filename).resolve()
    try:
        if OUTPUT_DIR.resolve() not in path.parents:
            return None
    except OSError:
        return None
    return path if path.exists() else None


def _recent_reports(limit: int = 10) -> list[Path]:
    if not OUTPUT_DIR.exists():
        return []
    files = [p for p in OUTPUT_DIR.glob("*.xlsx") if p.is_file()]
    return sorted(files, key=lambda p: p.stat().st_mtime, reverse=True)[:limit]


def _fmt_mtime(path: Path) -> str:
    return datetime.fromtimestamp(path.stat().st_mtime).strftime("%d.%m.%Y %H:%M")


def _read_env_file(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values
    for line in path.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value
    return values


def _current_settings() -> dict[str, str]:
    values = _read_env_file(ENV_EXAMPLE_PATH)
    values.update(_read_env_file(ENV_PATH))
    for key in SETTINGS_KEYS:
        if key not in values:
            values[key] = os.getenv(key, "")
    return values


def _setting_status(value: str) -> str:
    return "задано" if value else "не задано"


def _hash_password(password: str, salt: Optional[str] = None) -> str:
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        bytes.fromhex(salt),
        PASSWORD_HASH_ITERATIONS,
    ).hex()
    return f"pbkdf2_sha256${PASSWORD_HASH_ITERATIONS}${salt}${digest}"


def _verify_password(password: str, encoded: str) -> bool:
    try:
        algorithm, iterations_raw, salt, expected = encoded.split("$", 3)
        iterations = int(iterations_raw)
    except ValueError:
        return False
    if algorithm != "pbkdf2_sha256":
        return False
    digest = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        bytes.fromhex(salt),
        iterations,
    ).hex()
    return secrets.compare_digest(digest, expected)


def _env_admin_user() -> Optional[dict]:
    password = os.getenv("WEB_UI_PASSWORD", "").strip()
    if not password:
        return None
    username = os.getenv("WEB_UI_USER", "admin").strip() or "admin"
    return {
        "username": username,
        "password_hash": _hash_password(password),
        "is_admin": True,
        "blocked": False,
    }


def _load_users() -> dict[str, dict]:
    users: dict[str, dict] = {}
    if USERS_PATH.exists():
        try:
            data = json.loads(USERS_PATH.read_text(encoding="utf-8"))
            raw_users = data.get("users", {})
            if isinstance(raw_users, dict):
                for username, record in raw_users.items():
                    if isinstance(record, dict):
                        users[str(username)] = {
                            "username": str(username),
                            "password_hash": str(record.get("password_hash", "")),
                            "is_admin": bool(record.get("is_admin", False)),
                            "blocked": bool(record.get("blocked", False)),
                        }
        except (OSError, json.JSONDecodeError):
            users = {}
    if not users:
        admin = _env_admin_user()
        if admin:
            users[admin["username"]] = admin
    return users


def _save_users(users: dict[str, dict]) -> None:
    USERS_PATH.parent.mkdir(parents=True, exist_ok=True)
    data = {
        "users": {
            username: {
                "password_hash": record.get("password_hash", ""),
                "is_admin": bool(record.get("is_admin", False)),
                "blocked": bool(record.get("blocked", False)),
            }
            for username, record in sorted(users.items())
        }
    }
    tmp = USERS_PATH.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    os.chmod(tmp, 0o600)
    tmp.replace(USERS_PATH)
    os.chmod(USERS_PATH, 0o600)


def _authenticate_credentials(username: str, password: str) -> Optional[dict]:
    users = _load_users()
    if users:
        record = users.get(username)
        if not record or record.get("blocked"):
            return None
        if not _verify_password(password, record.get("password_hash", "")):
            return None
        return {
            "username": username,
            "is_admin": bool(record.get("is_admin", False)),
            "blocked": False,
        }

    expected_password = os.getenv("WEB_UI_PASSWORD", "").strip()
    if not expected_password:
        return {"username": "anonymous", "is_admin": True, "blocked": False}
    expected_user = os.getenv("WEB_UI_USER", "admin").strip() or "admin"
    if secrets.compare_digest(username, expected_user) and secrets.compare_digest(password, expected_password):
        return {"username": username, "is_admin": True, "blocked": False}
    return None


def _parse_basic_auth(header: str) -> Optional[tuple[str, str]]:
    if not header.startswith("Basic "):
        return None
    try:
        raw = base64.b64decode(header[6:]).decode("utf-8")
    except (ValueError, UnicodeDecodeError):
        return None
    username, sep, password = raw.partition(":")
    if not sep:
        return None
    return username, password


def _write_env_values(values: dict[str, str]) -> None:
    source = ENV_PATH if ENV_PATH.exists() else ENV_EXAMPLE_PATH
    lines = source.read_text(encoding="utf-8").splitlines() if source.exists() else []
    updated: list[str] = []
    seen: set[str] = set()

    for line in lines:
        stripped = line.strip()
        if stripped and not stripped.startswith("#") and "=" in line:
            key, _value = line.split("=", 1)
            key = key.strip()
            if key in values:
                updated.append(f"{key}={values[key]}")
                seen.add(key)
                continue
        updated.append(line)

    missing = [key for key in SETTINGS_KEYS if key not in seen]
    if missing:
        if updated and updated[-1].strip():
            updated.append("")
        for key in missing:
            updated.append(f"{key}={values.get(key, '')}")

    ENV_PATH.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(prefix=".env.", dir=str(ENV_PATH.parent), text=True)
    tmp_path = Path(tmp_name)
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as fh:
            fh.write("\n".join(updated).rstrip() + "\n")
        os.chmod(tmp_path, 0o600)
        tmp_path.replace(ENV_PATH)
        os.chmod(ENV_PATH, 0o600)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


def _reload_runtime_settings(values: dict[str, str]) -> None:
    for key in SETTINGS_KEYS:
        os.environ[key] = values.get(key, "")

    for module_name in ("config", "odata_client", "bitrix_sender", "email_sender"):
        module = sys.modules.get(module_name) or sys.modules.get(f"{__package__}.{module_name}")
        if module is None:
            continue
        if module_name == "config" and hasattr(module, "Settings"):
            module.settings = module.Settings.from_env()
        elif hasattr(module, "default_settings"):
            config_module = sys.modules.get("config") or sys.modules.get(f"{__package__}.config")
            if config_module is not None and hasattr(config_module, "settings"):
                module.default_settings = config_module.settings


def _validated_settings(form: dict[str, list[str]]) -> tuple[dict[str, str], Optional[str]]:
    current = _current_settings()
    values = dict(current)

    for field in SETTINGS_FIELDS:
        key = field["key"]
        kind = field.get("kind")
        if kind == "bool":
            values[key] = "1" if form.get(key, [""])[0] == "1" else "0"
            continue

        raw = form.get(key, [""])[0]
        if "\n" in raw or "\r" in raw:
            return values, f"{field['label']}: переносы строк недопустимы."
        raw = raw.strip()

        if field.get("secret"):
            if form.get(f"{key}__clear", [""])[0] == "1":
                values[key] = ""
            elif raw:
                values[key] = raw
            else:
                values[key] = current.get(key, "")
        else:
            values[key] = raw

    for field in SETTINGS_FIELDS:
        key = field["key"]
        value = values.get(key, "")
        kind = field.get("kind")
        if kind in ("int", "port"):
            try:
                numeric = int(value)
            except ValueError:
                return values, f"{field['label']}: укажите целое число."
            min_value = int(field.get("min", 1))
            max_value = int(field.get("max", 65535))
            if numeric < min_value or numeric > max_value:
                return values, f"{field['label']}: допустимо от {min_value} до {max_value}."
            if kind == "port" and numeric in (80, 443):
                return values, "Web UI не должен занимать порты 80 и 443."
            values[key] = str(numeric)
        elif kind == "time":
            if not value:
                return values, f"{field['label']}: укажите время."
            parts = value.split(":", 1)
            if len(parts) != 2 or not all(part.isdigit() for part in parts):
                return values, f"{field['label']}: время должно быть в формате HH:MM."
            hour, minute = int(parts[0]), int(parts[1])
            if hour > 23 or minute > 59:
                return values, f"{field['label']}: время вне диапазона 00:00..23:59."
            values[key] = f"{hour:02d}:{minute:02d}"
        elif kind == "select":
            allowed = {option[0] for option in field.get("options", [])}
            if value not in allowed:
                return values, f"{field['label']}: выберите значение из списка."

    return values, None


def _settings_form(values: dict[str, str]) -> str:
    sections: list[str] = []
    for section in dict.fromkeys(field["section"] for field in SETTINGS_FIELDS):
        controls = []
        for field in SETTINGS_FIELDS:
            if field["section"] != section:
                continue
            key = field["key"]
            value = values.get(key, "")
            label = html.escape(field["label"])
            help_text = field.get("help", "")
            autocomplete = html.escape(field.get("autocomplete", "off"))
            kind = field.get("kind")
            if kind == "bool":
                checked = " checked" if value.strip().lower() in ("1", "true", "yes", "on") else ""
                controls.append(
                    f"""
                    <label class="check">
                      <input type="checkbox" name="{key}" value="1"{checked}>
                      <span>{label}</span>
                    </label>
                    {f'<p class="hint">{html.escape(help_text)}</p>' if help_text else ''}
                    """
                )
            elif kind == "select":
                options = []
                for option_value, option_label in field.get("options", []):
                    selected = " selected" if option_value == value else ""
                    options.append(
                        f'<option value="{html.escape(option_value, quote=True)}"{selected}>'
                        f'{html.escape(option_label)}</option>'
                    )
                controls.append(
                    f"""
                    <div class="field">
                      <label class="title" for="{key}">{label}</label>
                      <select id="{key}" name="{key}">{''.join(options)}</select>
                      {f'<p class="hint">{html.escape(help_text)}</p>' if help_text else ''}
                    </div>
                    """
                )
            elif field.get("secret"):
                status = html.escape(_setting_status(value))
                controls.append(
                    f"""
                    <div class="field">
                      <div class="label-row">
                        <label class="title" for="{key}">{label}</label>
                        <span class="secret-state">{status}</span>
                      </div>
                      <input id="{key}" name="{key}" type="password"
                             autocomplete="{autocomplete}"
                             placeholder="Оставьте пустым, чтобы не менять">
                      <label class="check small">
                        <input type="checkbox" name="{key}__clear" value="1">
                        <span>Очистить значение</span>
                      </label>
                      {f'<p class="hint">{html.escape(help_text)}</p>' if help_text else ''}
                    </div>
                    """
                )
            else:
                input_type = "number" if kind in ("int", "port") else "time" if kind == "time" else "text"
                extra = ""
                if kind in ("int", "port"):
                    extra = (
                        f' min="{int(field.get("min", 1))}"'
                        f' max="{int(field.get("max", 65535))}"'
                        ' step="1"'
                    )
                controls.append(
                    f"""
                    <div class="field">
                      <label class="title" for="{key}">{label}</label>
                      <input id="{key}" name="{key}" type="{input_type}"
                             value="{html.escape(value, quote=True)}"
                             autocomplete="{autocomplete}"{extra}>
                      {f'<p class="hint">{html.escape(help_text)}</p>' if help_text else ''}
                    </div>
                    """
                )
        sections.append(
            f"""
            <fieldset>
              <legend>{html.escape(section)}</legend>
              {''.join(controls)}
            </fieldset>
            """
        )

    return f"""
    <section class="panel">
      <h2>Секреты и подключения</h2>
      <form method="post" action="/settings" autocomplete="off">
        {''.join(sections)}
        <div class="actions">
          <button type="submit">Сохранить настройки</button>
          <a class="button secondary" href="/">Вернуться к отчёту</a>
        </div>
      </form>
    </section>
    """


def _color_hex(color, default: str = "") -> str:
    if not color or color.type != "rgb" or not color.rgb:
        return default
    raw = str(color.rgb)
    if len(raw) == 8:
        raw = raw[2:]
    return f"#{raw}" if len(raw) == 6 else default


def _merged_map(ws) -> tuple[dict[tuple[int, int], tuple[int, int]], set[tuple[int, int]]]:
    tops: dict[tuple[int, int], tuple[int, int]] = {}
    covered: set[tuple[int, int]] = set()
    for merged in ws.merged_cells.ranges:
        tops[(merged.min_row, merged.min_col)] = (
            merged.max_row - merged.min_row + 1,
            merged.max_col - merged.min_col + 1,
        )
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                if (row, col) != (merged.min_row, merged.min_col):
                    covered.add((row, col))
    return tops, covered


def _numeric_cell(ws, coord: str, cache: dict[str, object]) -> Optional[float]:
    value = _display_cell_value(ws, coord, cache)
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def _display_cell_value(ws, coord: str, cache: dict[str, object]):
    if coord in cache:
        return cache[coord]
    value = ws[coord].value
    if not isinstance(value, str) or not value.startswith("="):
        cache[coord] = value
        return value

    formula = value.replace(" ", "")
    result = None

    sum_match = re.fullmatch(r"=SUM\(([A-Z]+\d+):([A-Z]+\d+)\)", formula)
    if sum_match:
        start, end = sum_match.groups()
        start_col = re.match(r"([A-Z]+)", start).group(1)
        start_row = int(re.search(r"(\d+)", start).group(1))
        end_row = int(re.search(r"(\d+)", end).group(1))
        total = 0.0
        for row in range(start_row, end_row + 1):
            total += _numeric_cell(ws, f"{start_col}{row}", cache) or 0.0
        result = total

    if result is None:
        math_match = re.fullmatch(r'=IFERROR\(([A-Z]+\d+)([-/])([A-Z]+\d+),"—"\)', formula)
        if math_match:
            left_ref, op, right_ref = math_match.groups()
            left = _numeric_cell(ws, left_ref, cache)
            right = _numeric_cell(ws, right_ref, cache)
            if left is not None and right is not None:
                if op == "-":
                    result = left - right
                elif right != 0:
                    result = left / right

    if result is None:
        ratio_match = re.fullmatch(
            r'=IFERROR\(IF\(AND\(ISNUMBER\(([A-Z]+\d+)\),ISNUMBER\(([A-Z]+\d+)\),'
            r'([A-Z]+\d+)>0\),([A-Z]+\d+)/([A-Z]+\d+),"—"\),"—"\)',
            formula,
        )
        if ratio_match:
            plan_ref, fact_ref, positive_ref, numerator_ref, denominator_ref = ratio_match.groups()
            plan = _numeric_cell(ws, plan_ref, cache)
            fact = _numeric_cell(ws, fact_ref, cache)
            positive = _numeric_cell(ws, positive_ref, cache)
            numerator = _numeric_cell(ws, numerator_ref, cache)
            denominator = _numeric_cell(ws, denominator_ref, cache)
            if (
                plan is not None
                and fact is not None
                and positive is not None
                and positive > 0
                and numerator is not None
                and denominator not in (None, 0)
            ):
                result = numerator / denominator

    cache[coord] = result if result is not None else "—"
    return cache[coord]


def _format_report_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if abs(value) >= 100:
            return f"{value:,.0f}".replace(",", " ")
        return f"{value:.2f}".rstrip("0").rstrip(".")
    return str(value)


def _cell_style(cell, row_height: Optional[float]) -> str:
    styles: list[str] = []
    fill = _color_hex(cell.fill.fgColor if cell.fill else None)
    if fill:
        styles.append(f"background:{fill}")
    color = _color_hex(cell.font.color if cell.font else None)
    if color:
        styles.append(f"color:{color}")
    if cell.font:
        if cell.font.bold:
            styles.append("font-weight:700")
        if cell.font.italic:
            styles.append("font-style:italic")
        if cell.font.sz:
            styles.append(f"font-size:{float(cell.font.sz):.1f}pt")
    if cell.alignment:
        if cell.alignment.horizontal:
            styles.append(f"text-align:{cell.alignment.horizontal}")
        if cell.alignment.vertical:
            styles.append(f"vertical-align:{cell.alignment.vertical}")
    if row_height:
        styles.append(f"height:{int(row_height * 1.35)}px")
    return ";".join(styles)


def _render_sheet_table(ws) -> str:
    tops, covered = _merged_map(ws)
    cache: dict[str, object] = {}
    colgroup = []
    for col in range(1, min(ws.max_column, 9) + 1):
        letter = get_column_letter(col)
        width = ws.column_dimensions[letter].width or 10
        colgroup.append(f'<col style="width:{max(18, int(width * 7.5))}px">')

    rows = []
    for row_idx in range(1, ws.max_row + 1):
        cells = []
        for col_idx in range(1, min(ws.max_column, 9) + 1):
            if (row_idx, col_idx) in covered:
                continue
            cell = ws.cell(row_idx, col_idx)
            rowspan, colspan = tops.get((row_idx, col_idx), (1, 1))
            attrs = []
            if rowspan > 1:
                attrs.append(f'rowspan="{rowspan}"')
            if colspan > 1:
                attrs.append(f'colspan="{colspan}"')
            style = _cell_style(cell, ws.row_dimensions[row_idx].height)
            if style:
                attrs.append(f'style="{html.escape(style, quote=True)}"')
            value = _display_cell_value(ws, cell.coordinate, cache)
            text = html.escape(_format_report_value(value))
            cells.append(f"<td {' '.join(attrs)}>{text}</td>")
        rows.append(f"<tr>{''.join(cells)}</tr>")
    return f"<table class=\"report-table\"><colgroup>{''.join(colgroup)}</colgroup>{''.join(rows)}</table>"


def _report_view_page(path: Path, sheet_name: str = "", message: str = "", error: str = "") -> str:
    wb = load_workbook(path, data_only=False)
    selected = sheet_name if sheet_name in wb.sheetnames else wb.sheetnames[0]
    ws = wb[selected]
    selected_index = wb.sheetnames.index(selected)
    prev_sheet = wb.sheetnames[selected_index - 1] if selected_index > 0 else ""
    next_sheet = wb.sheetnames[selected_index + 1] if selected_index < len(wb.sheetnames) - 1 else ""
    prev_button = (
        f'<a class="button" href="{html.escape(_view_url(path, prev_sheet), quote=True)}">Назад</a>'
        if prev_sheet
        else '<span class="button disabled">Назад</span>'
    )
    next_button = (
        f'<a class="button" href="{html.escape(_view_url(path, next_sheet), quote=True)}">Вперёд</a>'
        if next_sheet
        else '<span class="button disabled">Вперёд</span>'
    )
    message_html = f'<div class="notice ok">{html.escape(message)}</div>' if message else ""
    error_html = f'<div class="notice err">{html.escape(error)}</div>' if error else ""
    tabs = []
    for name in wb.sheetnames:
        cls = " active" if name == selected else ""
        tabs.append(
            f'<a class="sheet-tab{cls}" href="{html.escape(_view_url(path, name), quote=True)}">'
            f"{html.escape(name)}</a>"
        )
    table = _render_sheet_table(ws)
    return f"""<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{APP_TITLE} · {html.escape(selected)}</title>
  <style>
    :root {{
      --bg: #f6f7f8;
      --panel: #ffffff;
      --text: #202428;
      --muted: #687079;
      --line: #d9dee3;
      --accent: #167c80;
      --accent-dark: #0f5e61;
      --danger: #b42318;
      --ok: #1f7a4d;
      --shadow: 0 10px 28px rgba(32, 36, 40, .08);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      min-height: 100vh;
      background: var(--bg);
      color: var(--text);
      font: 15px/1.45 -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }}
    header {{
      border-bottom: 1px solid var(--line);
      background: #ffffff;
    }}
    .wrap {{
      width: min(1280px, calc(100% - 32px));
      margin: 0 auto;
    }}
    .topbar {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      min-height: 64px;
      gap: 16px;
    }}
    h1 {{
      margin: 0;
      font-size: 22px;
      font-weight: 650;
      letter-spacing: 0;
    }}
    nav {{
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
      align-items: center;
      justify-content: flex-end;
    }}
    nav a, .button {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-height: 34px;
      padding: 6px 10px;
      border-radius: 6px;
      color: var(--accent-dark);
      text-decoration: none;
      font-weight: 600;
      border: 1px solid transparent;
      background: transparent;
    }}
    nav a:hover, .button:hover {{ background: #eef3f3; }}
    .button.primary {{
      background: var(--accent);
      border-color: var(--accent);
      color: #fff;
    }}
    .button.primary:hover {{ background: var(--accent-dark); }}
    .button.disabled {{
      color: #9aa2aa;
      border-color: var(--line);
      cursor: default;
    }}
    .button.disabled:hover {{ background: transparent; }}
    main {{ padding: 22px 0 40px; }}
    .menu {{
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
      margin: 0 0 16px;
    }}
    .toolbar {{
      display: flex;
      gap: 10px;
      flex-wrap: wrap;
      align-items: center;
      justify-content: space-between;
      margin-bottom: 14px;
    }}
    .file-name {{
      color: var(--muted);
      font-size: 13px;
      overflow-wrap: anywhere;
    }}
    .nav-actions {{
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
      align-items: center;
      justify-content: flex-end;
    }}
    .notice {{
      border-radius: 8px;
      padding: 12px 14px;
      margin-bottom: 16px;
      border: 1px solid;
      background: #fff;
    }}
    .notice.ok {{
      color: var(--ok);
      border-color: rgba(31, 122, 77, .35);
    }}
    .notice.err {{
      color: var(--danger);
      border-color: rgba(180, 35, 24, .35);
    }}
    .tabs {{
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
      margin: 0 0 16px;
    }}
    .sheet-tab {{
      display: inline-flex;
      min-height: 34px;
      align-items: center;
      padding: 6px 10px;
      border: 1px solid var(--line);
      border-radius: 6px;
      color: var(--accent-dark);
      background: #fff;
      text-decoration: none;
      font-weight: 600;
    }}
    .sheet-tab.active {{
      background: var(--accent);
      border-color: var(--accent);
      color: #fff;
    }}
    .report-scroll {{
      overflow: auto;
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
      padding: 14px;
    }}
    .report-table {{
      border-collapse: collapse;
      table-layout: fixed;
      min-width: 980px;
      background: #fff;
    }}
    .report-table td {{
      border: 1px solid #b7c9d6;
      padding: 7px 8px;
      min-height: 24px;
      white-space: pre-wrap;
      overflow-wrap: anywhere;
      font-family: Arial, sans-serif;
      line-height: 1.28;
    }}
    @media (max-width: 760px) {{
      .topbar {{ align-items: flex-start; flex-direction: column; padding: 14px 0; }}
      nav {{ justify-content: flex-start; }}
      .toolbar {{ align-items: flex-start; flex-direction: column; }}
    }}
  </style>
</head>
<body>
  <header>
    <div class="wrap topbar">
      <h1>{APP_TITLE}</h1>
      <nav aria-label="Разделы">
        <a href="/">Создать отчёт</a>
        <a href="/">Файлы</a>
        <a href="/settings">Настройки</a>
        <a href="/users">Пользователи</a>
      </nav>
    </div>
  </header>
  <main class="wrap">
    {message_html}
    {error_html}
    <div class="menu">
      <a class="button primary" href="/">Новый отчёт</a>
      <a class="button" href="{html.escape(_file_url(path), quote=True)}">Скачать Excel</a>
      <a class="button" href="/settings">Настройки</a>
    </div>
    <div class="toolbar">
      <div>
        <strong>{html.escape(selected)}</strong>
        <div class="file-name">{html.escape(path.name)}</div>
      </div>
      <div class="nav-actions">
        {prev_button}
        <span class="file-name">{selected_index + 1} / {len(wb.sheetnames)}</span>
        {next_button}
      </div>
    </div>
    <div class="tabs">{''.join(tabs)}</div>
    <div class="report-scroll">{table}</div>
  </main>
</body>
</html>"""


def _page(
    *,
    selected_mode: str = "all",
    selected_date: Optional[date] = None,
    send_bitrix: bool = False,
    send_email: bool = False,
    message: str = "",
    error: str = "",
    generated: Optional[list[GeneratedReport]] = None,
) -> str:
    selected_date = selected_date or _default_date()
    mode_options = []
    for value, label in MODES.items():
        checked = " checked" if value == selected_mode else ""
        mode_options.append(
            f"""
            <label class="seg-item">
              <input type="radio" name="mode" value="{value}"{checked}>
              <span>{html.escape(label)}</span>
            </label>
            """
        )

    generated_html = ""
    if generated:
        items = []
        for report in generated:
            name = html.escape(report.path.name)
            items.append(
                f"""
                <li>
                  <span>{html.escape(report.kind)}</span>
                  <div class="file-row">
                    <a href="{_view_url(report.path)}">{name}</a>
                    <small><a href="{_view_url(report.path)}">Открыть</a> · <a href="{_file_url(report.path)}">Скачать</a></small>
                  </div>
                </li>
                """
            )
        generated_html = f"""
        <section class="panel">
          <h2>Готовый файл</h2>
          <ul class="file-list">{''.join(items)}</ul>
        </section>
        """

    recent_items = []
    for path in _recent_reports():
        recent_items.append(
            f"""
            <li>
              <span>{html.escape(_fmt_mtime(path))}</span>
              <div class="file-row">
                <a href="{_view_url(path)}">{html.escape(path.name)}</a>
                <small><a href="{_view_url(path)}">Открыть</a> · <a href="{_file_url(path)}">Скачать</a></small>
              </div>
            </li>
            """
        )
    recent_html = "".join(recent_items) or "<li><span></span><em>Нет файлов</em></li>"

    bitrix_checked = " checked" if send_bitrix else ""
    email_checked = " checked" if send_email else ""
    message_html = f'<div class="notice ok">{html.escape(message)}</div>' if message else ""
    error_html = f'<div class="notice err">{html.escape(error)}</div>' if error else ""

    return f"""<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{APP_TITLE}</title>
  <style>
    :root {{
      --bg: #f6f7f8;
      --panel: #ffffff;
      --text: #202428;
      --muted: #687079;
      --line: #d9dee3;
      --accent: #167c80;
      --accent-dark: #0f5e61;
      --danger: #b42318;
      --ok: #1f7a4d;
      --shadow: 0 10px 28px rgba(32, 36, 40, .08);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      min-height: 100vh;
      background: var(--bg);
      color: var(--text);
      font: 15px/1.45 -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }}
    header {{
      border-bottom: 1px solid var(--line);
      background: #ffffff;
    }}
    .wrap {{
      width: min(1040px, calc(100% - 32px));
      margin: 0 auto;
    }}
    .topbar {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      min-height: 64px;
      gap: 16px;
    }}
    h1 {{
      margin: 0;
      font-size: 22px;
      font-weight: 650;
      letter-spacing: 0;
    }}
    .status {{
      color: var(--muted);
      font-size: 13px;
      white-space: nowrap;
    }}
    main {{
      padding: 28px 0 40px;
    }}
    .grid {{
      display: grid;
      grid-template-columns: minmax(0, 1fr) minmax(320px, .8fr);
      gap: 20px;
      align-items: start;
    }}
    .panel {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
      padding: 20px;
    }}
    h2 {{
      margin: 0 0 16px;
      font-size: 16px;
      font-weight: 650;
      letter-spacing: 0;
    }}
    form {{
      display: grid;
      gap: 18px;
    }}
    .field {{
      display: grid;
      gap: 8px;
    }}
    label.title {{
      font-weight: 600;
      font-size: 13px;
      color: #30363c;
    }}
    input[type="date"], input[type="text"], input[type="password"], input[type="number"], input[type="time"], select {{
      width: 100%;
      min-height: 42px;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 8px 10px;
      font: inherit;
      color: var(--text);
      background: #fff;
    }}
    input::placeholder {{ color: #9aa2aa; }}
    .label-row {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 10px;
    }}
    .secret-state {{
      color: var(--muted);
      font-size: 12px;
      white-space: nowrap;
    }}
    .hint {{
      margin: -2px 0 0;
      color: var(--muted);
      font-size: 12px;
    }}
    fieldset {{
      display: grid;
      gap: 14px;
      margin: 0;
      padding: 0 0 18px;
      border: 0;
      border-bottom: 1px solid var(--line);
    }}
    fieldset:last-of-type {{
      border-bottom: 0;
      padding-bottom: 0;
    }}
    legend {{
      padding: 0;
      margin: 0 0 2px;
      font-size: 14px;
      font-weight: 700;
    }}
    .seg {{
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      border: 1px solid var(--line);
      border-radius: 8px;
      overflow: hidden;
      background: #eef1f3;
    }}
    .seg-item {{
      min-width: 0;
      cursor: pointer;
    }}
    .seg-item input {{
      position: absolute;
      opacity: 0;
      pointer-events: none;
    }}
    .seg-item span {{
      display: grid;
      place-items: center;
      min-height: 42px;
      padding: 8px 10px;
      color: var(--muted);
      border-right: 1px solid var(--line);
      text-align: center;
      white-space: nowrap;
    }}
    .seg-item:last-child span {{ border-right: 0; }}
    .seg-item input:checked + span {{
      background: #ffffff;
      color: var(--text);
      font-weight: 650;
    }}
    .check {{
      display: flex;
      align-items: center;
      gap: 10px;
      color: var(--text);
      font-weight: 500;
    }}
    .check input {{
      width: 18px;
      height: 18px;
      accent-color: var(--accent);
    }}
    .check.small {{
      gap: 8px;
      color: var(--muted);
      font-size: 13px;
      font-weight: 400;
    }}
    .check.small input {{
      width: 16px;
      height: 16px;
    }}
    .actions {{
      display: flex;
      gap: 12px;
      flex-wrap: wrap;
      align-items: center;
    }}
    button, .button {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-height: 42px;
      padding: 9px 16px;
      border-radius: 6px;
      border: 1px solid var(--accent);
      background: var(--accent);
      color: #fff;
      font: inherit;
      font-weight: 650;
      text-decoration: none;
      cursor: pointer;
    }}
    button:hover, .button:hover {{ background: var(--accent-dark); }}
    .button.secondary {{
      background: #fff;
      color: var(--accent-dark);
      border-color: var(--line);
    }}
    .button.secondary:hover {{
      background: #eef3f3;
      border-color: #b8c5c6;
    }}
    nav {{
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
      align-items: center;
      justify-content: flex-end;
    }}
    nav a {{
      display: inline-flex;
      align-items: center;
      min-height: 34px;
      padding: 6px 10px;
      border-radius: 6px;
      color: var(--accent-dark);
      text-decoration: none;
      font-weight: 600;
    }}
    nav a:hover {{ background: #eef3f3; }}
    .menu {{
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
      margin: 0 0 16px;
    }}
    .menu .button.primary {{
      background: var(--accent);
      border-color: var(--accent);
      color: #fff;
    }}
    .menu .button.primary:hover {{ background: var(--accent-dark); }}
    .notice {{
      border-radius: 8px;
      padding: 12px 14px;
      margin-bottom: 16px;
      border: 1px solid;
      background: #fff;
    }}
    .notice.ok {{
      color: var(--ok);
      border-color: rgba(31, 122, 77, .35);
    }}
    .notice.err {{
      color: var(--danger);
      border-color: rgba(180, 35, 24, .35);
    }}
    .file-list {{
      list-style: none;
      padding: 0;
      margin: 0;
      display: grid;
      gap: 10px;
    }}
    .file-list li {{
      display: grid;
      grid-template-columns: 140px minmax(0, 1fr);
      gap: 12px;
      align-items: baseline;
      border-bottom: 1px solid var(--line);
      padding-bottom: 10px;
    }}
    .file-list li:last-child {{
      border-bottom: 0;
      padding-bottom: 0;
    }}
    .file-list span, .file-list em {{
      color: var(--muted);
      font-size: 13px;
      font-style: normal;
    }}
    .file-row {{
      display: grid;
      gap: 3px;
      min-width: 0;
    }}
    .file-row small {{
      color: var(--muted);
      font-size: 12px;
    }}
    a {{
      color: var(--accent-dark);
      overflow-wrap: anywhere;
      text-decoration-thickness: 1px;
      text-underline-offset: 3px;
    }}
    @media (max-width: 760px) {{
      .grid {{ grid-template-columns: 1fr; }}
      .topbar {{ align-items: flex-start; flex-direction: column; padding: 14px 0; }}
      .status {{ white-space: normal; }}
      .file-list li {{ grid-template-columns: 1fr; gap: 4px; }}
      .seg {{ grid-template-columns: 1fr; }}
      .seg-item span {{ border-right: 0; border-bottom: 1px solid var(--line); }}
      .seg-item:last-child span {{ border-bottom: 0; }}
    }}
  </style>
</head>
<body>
  <header>
    <div class="wrap topbar">
      <h1>{APP_TITLE}</h1>
      <nav aria-label="Разделы">
        <a href="/">Создать отчёт</a>
        <a href="#files">Файлы</a>
        <a href="/settings">Настройки</a>
        <a href="/users">Пользователи</a>
      </nav>
    </div>
  </header>
  <main class="wrap">
    {message_html}
    {error_html}
    <div class="menu">
      <a class="button primary" href="/">Новый отчёт</a>
      <a class="button" href="#files">Последние файлы</a>
      <a class="button" href="/settings">Настройки</a>
      <a class="button" href="/users">Пользователи</a>
    </div>
    <div class="grid">
      <section class="panel">
        <h2>Настройка отчёта</h2>
        <form method="post" action="/generate">
          <div class="field">
            <label class="title">Режим</label>
            <div class="seg">{''.join(mode_options)}</div>
          </div>
          <div class="field">
            <label class="title" for="date">Дата</label>
            <input id="date" name="date" type="date" value="{selected_date.isoformat()}" required>
          </div>
          <label class="check">
            <input type="checkbox" name="send_bitrix" value="1"{bitrix_checked}>
            <span>Отправить в Bitrix24</span>
          </label>
          <label class="check">
            <input type="checkbox" name="send_email" value="1"{email_checked}>
            <span>Отправить по email</span>
          </label>
          <div class="actions">
            <button type="submit">Сформировать</button>
          </div>
        </form>
      </section>
      <section id="files" class="panel">
        <h2>Последние файлы</h2>
        <ul class="file-list">{recent_html}</ul>
      </section>
      {generated_html}
    </div>
  </main>
</body>
</html>"""


def _settings_page(message: str = "", error: str = "") -> str:
    message_html = f'<div class="notice ok">{html.escape(message)}</div>' if message else ""
    error_html = f'<div class="notice err">{html.escape(error)}</div>' if error else ""
    return f"""<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{APP_TITLE} · Настройки</title>
  <style>
    :root {{
      --bg: #f6f7f8;
      --panel: #ffffff;
      --text: #202428;
      --muted: #687079;
      --line: #d9dee3;
      --accent: #167c80;
      --accent-dark: #0f5e61;
      --danger: #b42318;
      --ok: #1f7a4d;
      --shadow: 0 10px 28px rgba(32, 36, 40, .08);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      min-height: 100vh;
      background: var(--bg);
      color: var(--text);
      font: 15px/1.45 -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }}
    header {{
      border-bottom: 1px solid var(--line);
      background: #ffffff;
    }}
    .wrap {{
      width: min(1040px, calc(100% - 32px));
      margin: 0 auto;
    }}
    .topbar {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      min-height: 64px;
      gap: 16px;
    }}
    h1 {{
      margin: 0;
      font-size: 22px;
      font-weight: 650;
      letter-spacing: 0;
    }}
    main {{
      padding: 28px 0 40px;
    }}
    .panel {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
      padding: 20px;
    }}
    h2 {{
      margin: 0 0 16px;
      font-size: 16px;
      font-weight: 650;
      letter-spacing: 0;
    }}
    form {{
      display: grid;
      gap: 18px;
    }}
    .field {{
      display: grid;
      gap: 8px;
    }}
    label.title {{
      font-weight: 600;
      font-size: 13px;
      color: #30363c;
    }}
    input[type="date"], input[type="text"], input[type="password"], input[type="number"], input[type="time"], select {{
      width: 100%;
      min-height: 42px;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 8px 10px;
      font: inherit;
      color: var(--text);
      background: #fff;
    }}
    input::placeholder {{ color: #9aa2aa; }}
    .label-row {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 10px;
    }}
    .secret-state {{
      color: var(--muted);
      font-size: 12px;
      white-space: nowrap;
    }}
    .hint {{
      margin: -2px 0 0;
      color: var(--muted);
      font-size: 12px;
    }}
    fieldset {{
      display: grid;
      gap: 14px;
      margin: 0;
      padding: 0 0 18px;
      border: 0;
      border-bottom: 1px solid var(--line);
    }}
    fieldset:last-of-type {{
      border-bottom: 0;
      padding-bottom: 0;
    }}
    legend {{
      padding: 0;
      margin: 0 0 2px;
      font-size: 14px;
      font-weight: 700;
    }}
    .check {{
      display: flex;
      align-items: center;
      gap: 10px;
      color: var(--text);
      font-weight: 500;
    }}
    .check input {{
      width: 18px;
      height: 18px;
      accent-color: var(--accent);
    }}
    .check.small {{
      gap: 8px;
      color: var(--muted);
      font-size: 13px;
      font-weight: 400;
    }}
    .check.small input {{
      width: 16px;
      height: 16px;
    }}
    .actions {{
      display: flex;
      gap: 12px;
      flex-wrap: wrap;
      align-items: center;
    }}
    button, .button {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-height: 42px;
      padding: 9px 16px;
      border-radius: 6px;
      border: 1px solid var(--accent);
      background: var(--accent);
      color: #fff;
      font: inherit;
      font-weight: 650;
      text-decoration: none;
      cursor: pointer;
    }}
    button:hover, .button:hover {{ background: var(--accent-dark); }}
    .button.secondary {{
      background: #fff;
      color: var(--accent-dark);
      border-color: var(--line);
    }}
    .button.secondary:hover {{
      background: #eef3f3;
      border-color: #b8c5c6;
    }}
    nav {{
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
      align-items: center;
      justify-content: flex-end;
    }}
    nav a {{
      display: inline-flex;
      align-items: center;
      min-height: 34px;
      padding: 6px 10px;
      border-radius: 6px;
      color: var(--accent-dark);
      text-decoration: none;
      font-weight: 600;
    }}
    nav a:hover {{ background: #eef3f3; }}
    .notice {{
      border-radius: 8px;
      padding: 12px 14px;
      margin-bottom: 16px;
      border: 1px solid;
      background: #fff;
    }}
    .notice.ok {{
      color: var(--ok);
      border-color: rgba(31, 122, 77, .35);
    }}
    .notice.err {{
      color: var(--danger);
      border-color: rgba(180, 35, 24, .35);
    }}
    @media (max-width: 760px) {{
      .topbar {{ align-items: flex-start; flex-direction: column; padding: 14px 0; }}
      nav {{ justify-content: flex-start; }}
    }}
  </style>
</head>
<body>
  <header>
    <div class="wrap topbar">
      <h1>{APP_TITLE}</h1>
      <nav aria-label="Разделы">
        <a href="/">Создать отчёт</a>
        <a href="/settings">Настройки</a>
        <a href="/users">Пользователи</a>
      </nav>
    </div>
  </header>
  <main class="wrap">
    {message_html}
    {error_html}
    {_settings_form(_current_settings())}
  </main>
</body>
</html>"""


def _active_admin_count(users: dict[str, dict]) -> int:
    return sum(1 for record in users.values() if record.get("is_admin") and not record.get("blocked"))


def _users_page(current_user: dict, message: str = "", error: str = "") -> str:
    users = _load_users()
    message_html = f'<div class="notice ok">{html.escape(message)}</div>' if message else ""
    error_html = f'<div class="notice err">{html.escape(error)}</div>' if error else ""
    rows = []
    for username, record in sorted(users.items()):
        is_self = username == current_user.get("username")
        blocked = bool(record.get("blocked"))
        role = "admin" if record.get("is_admin") else "user"
        status = "заблокирован" if blocked else "активен"
        block_action = "unblock" if blocked else "block"
        block_label = "Разблокировать" if blocked else "Заблокировать"
        disabled_self = " disabled" if is_self else ""
        rows.append(
            f"""
            <tr>
              <td><strong>{html.escape(username)}</strong>{' <span class="muted">(вы)</span>' if is_self else ''}</td>
              <td>{html.escape(role)}</td>
              <td>{html.escape(status)}</td>
              <td>
                <form method="post" action="/users" class="inline-form">
                  <input type="hidden" name="action" value="password">
                  <input type="hidden" name="username" value="{html.escape(username, quote=True)}">
                  <input type="password" name="password" placeholder="Новый пароль" autocomplete="new-password" required>
                  <button type="submit">Сменить</button>
                </form>
              </td>
              <td>
                <form method="post" action="/users" class="inline-form">
                  <input type="hidden" name="action" value="{block_action}">
                  <input type="hidden" name="username" value="{html.escape(username, quote=True)}">
                  <button type="submit"{disabled_self}>{block_label}</button>
                </form>
              </td>
              <td>
                <form method="post" action="/users" class="inline-form">
                  <input type="hidden" name="action" value="delete">
                  <input type="hidden" name="username" value="{html.escape(username, quote=True)}">
                  <button type="submit"{disabled_self}>Удалить</button>
                </form>
              </td>
            </tr>
            """
        )
    table = "".join(rows) or '<tr><td colspan="6">Пользователей нет</td></tr>'
    return f"""<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{APP_TITLE} · Пользователи</title>
  <style>
    :root {{
      --bg: #f6f7f8;
      --panel: #ffffff;
      --text: #202428;
      --muted: #687079;
      --line: #d9dee3;
      --accent: #167c80;
      --accent-dark: #0f5e61;
      --danger: #b42318;
      --ok: #1f7a4d;
      --shadow: 0 10px 28px rgba(32, 36, 40, .08);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      min-height: 100vh;
      background: var(--bg);
      color: var(--text);
      font: 15px/1.45 -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }}
    header {{ border-bottom: 1px solid var(--line); background: #fff; }}
    .wrap {{ width: min(1180px, calc(100% - 32px)); margin: 0 auto; }}
    .topbar {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      min-height: 64px;
      gap: 16px;
    }}
    h1 {{ margin: 0; font-size: 22px; font-weight: 650; letter-spacing: 0; }}
    main {{ padding: 28px 0 40px; }}
    nav {{ display: flex; gap: 8px; flex-wrap: wrap; align-items: center; justify-content: flex-end; }}
    nav a, .button {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-height: 34px;
      padding: 6px 10px;
      border-radius: 6px;
      color: var(--accent-dark);
      text-decoration: none;
      font-weight: 600;
      border: 1px solid transparent;
      background: transparent;
    }}
    nav a:hover, .button:hover {{ background: #eef3f3; }}
    .panel {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
      padding: 20px;
      margin-bottom: 18px;
    }}
    h2 {{ margin: 0 0 16px; font-size: 16px; font-weight: 650; letter-spacing: 0; }}
    input[type="text"], input[type="password"] {{
      min-height: 38px;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 7px 10px;
      font: inherit;
      color: var(--text);
      background: #fff;
    }}
    button {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-height: 38px;
      padding: 7px 12px;
      border-radius: 6px;
      border: 1px solid var(--accent);
      background: var(--accent);
      color: #fff;
      font: inherit;
      font-weight: 650;
      cursor: pointer;
    }}
    button:hover {{ background: var(--accent-dark); }}
    button:disabled {{
      cursor: not-allowed;
      border-color: var(--line);
      background: #eef1f3;
      color: var(--muted);
    }}
    .notice {{
      border-radius: 8px;
      padding: 12px 14px;
      margin-bottom: 16px;
      border: 1px solid;
      background: #fff;
    }}
    .notice.ok {{ color: var(--ok); border-color: rgba(31, 122, 77, .35); }}
    .notice.err {{ color: var(--danger); border-color: rgba(180, 35, 24, .35); }}
    .create-form {{
      display: grid;
      grid-template-columns: minmax(180px, 1fr) minmax(180px, 1fr) auto auto;
      gap: 10px;
      align-items: center;
    }}
    .hint {{
      margin: -6px 0 14px;
      color: var(--muted);
      font-size: 13px;
    }}
    .check {{ display: flex; align-items: center; gap: 8px; color: var(--text); font-weight: 500; }}
    .table-wrap {{ overflow: auto; }}
    table {{ width: 100%; border-collapse: collapse; min-width: 900px; }}
    th, td {{ border-bottom: 1px solid var(--line); padding: 10px; text-align: left; vertical-align: middle; }}
    th {{ color: #30363c; font-size: 13px; background: #eef3f3; }}
    .inline-form {{ display: flex; gap: 8px; flex-wrap: wrap; align-items: center; margin: 0; }}
    .inline-form input[type="password"] {{ width: 180px; }}
    .muted {{ color: var(--muted); font-size: 12px; }}
    @media (max-width: 760px) {{
      .topbar {{ align-items: flex-start; flex-direction: column; padding: 14px 0; }}
      nav {{ justify-content: flex-start; }}
      .create-form {{ grid-template-columns: 1fr; }}
    }}
  </style>
</head>
<body>
  <header>
    <div class="wrap topbar">
      <h1>{APP_TITLE}</h1>
      <nav aria-label="Разделы">
        <a href="/">Создать отчёт</a>
        <a href="/settings">Настройки</a>
        <a href="/users">Пользователи</a>
      </nav>
    </div>
  </header>
  <main class="wrap">
    {message_html}
    {error_html}
    <section class="panel">
      <h2>Новый пользователь</h2>
      <p class="hint">Учётные записи сохраняются в backend/users.json. Управление доступно только роли admin.</p>
      <form method="post" action="/users" class="create-form">
        <input type="hidden" name="action" value="create">
        <input type="text" name="username" placeholder="Логин" autocomplete="username" required>
        <input type="password" name="password" placeholder="Пароль" autocomplete="new-password" required>
        <label class="check"><input type="checkbox" name="is_admin" value="1"> <span>admin</span></label>
        <button type="submit">Создать</button>
      </form>
    </section>
    <section class="panel">
      <h2>Пользователи</h2>
      <p class="hint">Текущего пользователя и последнего активного admin нельзя удалить или заблокировать.</p>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>Логин</th>
              <th>Роль</th>
              <th>Статус</th>
              <th>Пароль</th>
              <th>Доступ</th>
              <th>Удаление</th>
            </tr>
          </thead>
          <tbody>{table}</tbody>
        </table>
      </div>
    </section>
  </main>
</body>
</html>"""


def _apply_schedule(values: dict[str, str]) -> str:
    try:
        entries = install_schedule(values)
    except (ScheduleError, OSError) as exc:
        return f"Расписание не обновлено: {exc}"
    if entries:
        return f"Расписание обновлено: {len(entries)} задач."
    return "Расписание выключено, задачи cron удалены."


def _reports_from_query(params: dict[str, list[str]]) -> list[GeneratedReport]:
    reports: list[GeneratedReport] = []
    for filename in params.get("file", []):
        path = _safe_output_path(filename)
        if path is not None:
            reports.append(GeneratedReport("Файл", filename, path))
    return reports


class ReportWebHandler(BaseHTTPRequestHandler):
    server_version = "ATCReportWeb/1.0"

    def _auth_required(self) -> bool:
        return bool(_load_users() or os.getenv("WEB_UI_PASSWORD", "").strip())

    def _current_user(self) -> Optional[dict]:
        if hasattr(self, "_cached_user"):
            return self._cached_user
        if not self._auth_required():
            self._cached_user = {"username": "anonymous", "is_admin": True, "blocked": False}
            return self._cached_user
        credentials = _parse_basic_auth(self.headers.get("Authorization", ""))
        if credentials is None:
            self._cached_user = None
            return None
        self._cached_user = _authenticate_credentials(credentials[0], credentials[1])
        return self._cached_user

    def _is_authorized(self) -> bool:
        return self._current_user() is not None

    def _require_auth(self) -> bool:
        if self._is_authorized():
            return False
        self.send_response(HTTPStatus.UNAUTHORIZED)
        self.send_header("WWW-Authenticate", 'Basic realm="ATC Reports"')
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.end_headers()
        self.wfile.write("Требуется авторизация".encode("utf-8"))
        return True

    def _require_admin(self) -> bool:
        if self._require_auth():
            return True
        user = self._current_user()
        if user and user.get("is_admin"):
            return False
        self.send_response(HTTPStatus.FORBIDDEN)
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.end_headers()
        self.wfile.write("Доступно только admin".encode("utf-8"))
        return True

    def _send_html(self, content: str, status: HTTPStatus = HTTPStatus.OK) -> None:
        data = content.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def _redirect(self, location: str) -> None:
        self.send_response(HTTPStatus.SEE_OTHER)
        self.send_header("Location", location)
        self.send_header("Content-Length", "0")
        self.end_headers()

    def _report_location(
        self,
        *,
        mode: str = "all",
        report_date: Optional[date] = None,
        send_bitrix: bool = False,
        send_email: bool = False,
        message: str = "",
        error: str = "",
        generated: Optional[list[GeneratedReport]] = None,
    ) -> str:
        query: dict[str, object] = {
            "mode": mode,
            "send_bitrix": "1" if send_bitrix else "0",
            "send_email": "1" if send_email else "0",
        }
        if report_date is not None:
            query["date"] = report_date.isoformat()
        if message:
            query["message"] = message
        if error:
            query["error"] = error
        if generated:
            query["file"] = [report.path.name for report in generated]
        return "/?" + urlencode(query, doseq=True)

    def _settings_location(self, *, message: str = "", error: str = "") -> str:
        query = {}
        if message:
            query["message"] = message
        if error:
            query["error"] = error
        return "/settings" + (("?" + urlencode(query)) if query else "")

    def _users_location(self, *, message: str = "", error: str = "") -> str:
        query = {}
        if message:
            query["message"] = message
        if error:
            query["error"] = error
        return "/users" + (("?" + urlencode(query)) if query else "")

    def _send_error_page(
        self,
        error: str,
        *,
        selected_mode: str = "all",
        selected_date: Optional[date] = None,
        send_bitrix: bool = False,
    ) -> None:
        self._send_html(
            _page(
                selected_mode=selected_mode,
                selected_date=selected_date,
                send_bitrix=send_bitrix,
                send_email=False,
                error=error,
            ),
            HTTPStatus.BAD_REQUEST,
        )

    def do_GET(self) -> None:  # noqa: N802
        if self._require_auth():
            return
        parsed = urlparse(self.path)
        if parsed.path == "/":
            params = parse_qs(parsed.query)
            mode = params.get("mode", ["all"])[0]
            if mode not in MODES:
                mode = "all"
            try:
                selected_date = _parse_report_date(params.get("date", [""])[0])
            except ValueError:
                selected_date = _default_date()
            self._send_html(
                _page(
                    selected_mode=mode,
                    selected_date=selected_date,
                    send_bitrix=params.get("send_bitrix", [""])[0] == "1",
                    send_email=params.get("send_email", [""])[0] == "1",
                    message=params.get("message", [""])[0],
                    error=params.get("error", [""])[0],
                    generated=_reports_from_query(params),
                )
            )
            return
        if parsed.path == "/settings":
            params = parse_qs(parsed.query)
            self._send_html(
                _settings_page(
                    message=params.get("message", [""])[0],
                    error=params.get("error", [""])[0],
                )
            )
            return
        if parsed.path == "/users":
            if self._require_admin():
                return
            params = parse_qs(parsed.query)
            self._send_html(
                _users_page(
                    self._current_user() or {},
                    message=params.get("message", [""])[0],
                    error=params.get("error", [""])[0],
                )
            )
            return
        if parsed.path == "/view":
            self._view_report(parsed.query)
            return
        if parsed.path == "/download":
            self._download(parsed.query)
            return
        if parsed.path == "/health":
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "text/plain; charset=utf-8")
            self.end_headers()
            self.wfile.write("ok".encode("utf-8"))
            return
        self.send_error(HTTPStatus.NOT_FOUND)

    def do_HEAD(self) -> None:  # noqa: N802
        if self._require_auth():
            return
        parsed = urlparse(self.path)
        if parsed.path == "/":
            data = _page().encode("utf-8")
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            return
        if parsed.path == "/settings":
            data = _settings_page().encode("utf-8")
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            return
        if parsed.path == "/users":
            if self._require_admin():
                return
            data = _users_page(self._current_user() or {}).encode("utf-8")
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            return
        if parsed.path == "/view":
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            return
        if parsed.path == "/health":
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "text/plain; charset=utf-8")
            self.send_header("Content-Length", "2")
            self.end_headers()
            return
        self.send_error(HTTPStatus.NOT_FOUND)

    def do_POST(self) -> None:  # noqa: N802
        if self._require_auth():
            return
        parsed = urlparse(self.path)
        if parsed.path == "/settings":
            self._save_settings()
            return
        if parsed.path == "/users":
            self._save_user_action()
            return
        if parsed.path == "/generate":
            self._generate_report()
            return
        self.send_error(HTTPStatus.NOT_FOUND)

    def _save_user_action(self) -> None:
        if self._require_admin():
            return
        current = self._current_user() or {}
        length = int(self.headers.get("Content-Length", "0") or "0")
        raw = self.rfile.read(length).decode("utf-8")
        form = parse_qs(raw)
        action = form.get("action", [""])[0]
        username = form.get("username", [""])[0].strip()
        users = _load_users()

        if not username or not re.fullmatch(r"[A-Za-z0-9_.@-]{3,64}", username):
            self._redirect(self._users_location(error="Логин должен быть 3-64 символа: латиница, цифры, _ . @ -"))
            return

        try:
            if action == "create":
                if username in users:
                    raise ValueError("Пользователь уже существует")
                password = form.get("password", [""])[0]
                if len(password) < 8:
                    raise ValueError("Пароль должен быть не короче 8 символов")
                users[username] = {
                    "username": username,
                    "password_hash": _hash_password(password),
                    "is_admin": form.get("is_admin", [""])[0] == "1",
                    "blocked": False,
                }
                message = "Пользователь создан."
            elif action == "password":
                if username not in users:
                    raise ValueError("Пользователь не найден")
                password = form.get("password", [""])[0]
                if len(password) < 8:
                    raise ValueError("Пароль должен быть не короче 8 символов")
                users[username]["password_hash"] = _hash_password(password)
                message = "Пароль изменён."
            elif action in ("block", "unblock"):
                if username not in users:
                    raise ValueError("Пользователь не найден")
                if username == current.get("username"):
                    raise ValueError("Нельзя заблокировать самого себя")
                new_blocked = action == "block"
                if new_blocked and users[username].get("is_admin") and _active_admin_count(users) <= 1:
                    raise ValueError("Нельзя заблокировать последнего активного admin")
                users[username]["blocked"] = new_blocked
                message = "Доступ изменён."
            elif action == "delete":
                if username not in users:
                    raise ValueError("Пользователь не найден")
                if username == current.get("username"):
                    raise ValueError("Нельзя удалить самого себя")
                if users[username].get("is_admin") and _active_admin_count(users) <= 1:
                    raise ValueError("Нельзя удалить последнего активного admin")
                del users[username]
                message = "Пользователь удалён."
            else:
                raise ValueError("Неизвестное действие")
            _save_users(users)
        except (OSError, ValueError) as exc:
            self._redirect(self._users_location(error=str(exc)))
            return
        self._redirect(self._users_location(message=message))

    def _save_settings(self) -> None:
        length = int(self.headers.get("Content-Length", "0") or "0")
        raw = self.rfile.read(length).decode("utf-8")
        form = parse_qs(raw)
        values, error = _validated_settings(form)
        if error:
            self._redirect(self._settings_location(error=error))
            return
        try:
            _write_env_values(values)
            _reload_runtime_settings(values)
        except OSError as exc:
            self._redirect(self._settings_location(error=f"Не удалось сохранить backend/.env: {exc}"))
            return
        schedule_message = _apply_schedule(values)
        self._redirect(self._settings_location(message=f"Настройки сохранены. {schedule_message}"))

    def _generate_report(self) -> None:
        length = int(self.headers.get("Content-Length", "0") or "0")
        raw = self.rfile.read(length).decode("utf-8")
        form = parse_qs(raw)
        mode = form.get("mode", ["all"])[0]
        date_raw = form.get("date", [""])[0]
        send_bitrix = form.get("send_bitrix", [""])[0] == "1"
        send_email = form.get("send_email", [""])[0] == "1"

        if mode not in MODES:
            self._redirect(
                self._report_location(
                    mode="all",
                    send_bitrix=send_bitrix,
                    send_email=send_email,
                    error="Неизвестный режим отчёта.",
                )
            )
            return
        try:
            report_date = _parse_report_date(date_raw)
        except ValueError:
            self._redirect(
                self._report_location(
                    mode=mode,
                    send_bitrix=send_bitrix,
                    send_email=send_email,
                    error="Дата должна быть в формате YYYY-MM-DD.",
                )
            )
            return

        started = time.monotonic()
        try:
            service = MetricsService()
            reporter = ExcelReporter()
            daily_day = report_date
            weekly_day = report_date
            if mode == "weekly":
                weekly_day = report_date
            generated = generate_workbook(service, reporter, mode, daily_day, weekly_day)
            if send_bitrix:
                send_to_bitrix(generated, "Отчёты АТЦ: Excel-файл с листами отчётов")
            if send_email:
                send_to_email(generated, "Отчёты АТЦ: Excel-файл с листами отчётов")
        except ODataUnavailableError as exc:
            self._redirect(
                self._report_location(
                    mode=mode,
                    report_date=report_date,
                    send_bitrix=send_bitrix,
                    send_email=send_email,
                    error=f"OData недоступен: {exc}",
                )
            )
            return
        except (ODataError, BitrixError, EmailError, RuntimeError) as exc:
            self._redirect(
                self._report_location(
                    mode=mode,
                    report_date=report_date,
                    send_bitrix=send_bitrix,
                    send_email=send_email,
                    error=str(exc),
                )
            )
            return

        elapsed = time.monotonic() - started
        channels = []
        if send_bitrix:
            channels.append("Bitrix24")
        if send_email:
            channels.append("email")
        suffix = f" Отправлено: {', '.join(channels)}." if channels else ""
        self._redirect(
            _view_location(
                generated[0].path,
                message=f"Отчёт сформирован за {elapsed:.1f} сек.{suffix}",
            )
        )

    def _download(self, query: str) -> None:
        params = parse_qs(query)
        filename = unquote(params.get("file", [""])[0])
        path = _safe_output_path(filename)
        if path is None:
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        data = path.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header(
            "Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.send_header("Content-Length", str(len(data)))
        self.send_header(
            "Content-Disposition",
            f"attachment; filename*=UTF-8''{quote(path.name)}",
        )
        self.end_headers()
        self.wfile.write(data)

    def _view_report(self, query: str) -> None:
        params = parse_qs(query)
        filename = unquote(params.get("file", [""])[0])
        path = _safe_output_path(filename)
        if path is None:
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        sheet_name = params.get("sheet", [""])[0]
        try:
            self._send_html(
                _report_view_page(
                    path,
                    sheet_name,
                    message=params.get("message", [""])[0],
                    error=params.get("error", [""])[0],
                )
            )
        except (OSError, ValueError) as exc:
            self._send_html(
                _page(error=f"Не удалось открыть книгу для просмотра: {exc}"),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    def log_message(self, fmt: str, *args) -> None:
        sys.stderr.write("[%s] %s\n" % (self.log_date_time_string(), fmt % args))


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Web UI настройки отчётов АТЦ")
    parser.add_argument(
        "--host",
        default=os.getenv("WEB_UI_HOST", "127.0.0.1").strip() or "127.0.0.1",
        help="адрес для web-интерфейса",
    )
    parser.add_argument(
        "--port",
        type=int,
        default=_env_int("WEB_UI_PORT", 8080),
        help="порт для web-интерфейса",
    )
    return parser


def main(argv: Optional[list[str]] = None) -> int:
    args = build_parser().parse_args(argv)
    host = args.host
    port = args.port
    if port in (80, 443):
        print("WEB UI must not use ports 80 or 443. Set WEB_UI_PORT to 8080 or another high port.", file=sys.stderr)
        return 2
    password_set = bool(os.getenv("WEB_UI_PASSWORD", "").strip())
    if host not in ("127.0.0.1", "localhost") and not password_set:
        print("Warning: WEB_UI_PASSWORD is empty while WEB_UI_HOST is not local.", file=sys.stderr)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    server = ThreadingHTTPServer((host, port), ReportWebHandler)
    print(f"Web UI: http://{host}:{port}/")
    if password_set:
        print(f"Basic Auth user: {os.getenv('WEB_UI_USER', 'admin') or 'admin'}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopped")
    finally:
        server.server_close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
