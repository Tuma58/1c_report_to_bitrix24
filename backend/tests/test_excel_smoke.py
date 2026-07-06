"""Smoke-тест ExcelReporter (Инкремент 3).

Считает MetricsService.daily() для «Арсенал» и «Реф. Сервис» за дату с данными,
заполняет копию шаблона и сохраняет в output/. Затем открывает сохранённые
файлы через openpyxl и проверяет:
- C3/C4 заполнены;
- D8/F8, D9/F9, D10/F10, D14/F14, D15/F15, D19/F19, D23 содержат значения;
- формульные ячейки (D16/D17/D18/H8) остались формулами (начинаются с '=').

Только GET из 1С; запись только в выходной xlsx. Если OData недоступен —
это не вина кода (выход с кодом 2).
"""
from __future__ import annotations

import os
import sys
from datetime import date

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

from openpyxl import load_workbook  # noqa: E402

from odata_client import ODataUnavailableError  # noqa: E402
from metrics import MetricsService  # noqa: E402
from excel_reporter import ExcelReporter  # noqa: E402


TEST_DAY = date(2026, 6, 10)  # будний день июня 2026 — есть ФАКТ
REPORTS = ["Арсенал", "Реф. Сервис"]

_HERE = os.path.dirname(__file__)
OUTPUT_DIR = os.path.abspath(os.path.join(_HERE, "..", "output"))

# Ячейки со значениями (должны быть заполнены, тип число или прочерк).
VALUE_CELLS = ["C3", "C4", "D8", "F8", "D9", "F9", "D10", "F10",
               "D14", "F14", "D15", "F15", "D19", "F19", "D23"]
# Формульные ячейки — должны остаться формулами.
FORMULA_CELLS = ["D16", "F16", "D17", "F17", "D18", "F18", "H8"]
SHEET_BY_REPORT = {"Арсенал": "Арсенал (Д)", "Реф. Сервис": "Реф.сервис (Д)"}


def _safe_name(report_name: str) -> str:
    return report_name.replace(" ", "_").replace(".", "").replace("/", "_")


def _is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


def verify_file(path: str, sheet: str) -> list[str]:
    """Открывает файл и возвращает список нарушений (пусто = OK)."""
    problems: list[str] = []
    wb = load_workbook(path, data_only=False)
    ws = wb[sheet]

    print(f"\n  Проверка файла: {path}")
    print(f"  Лист: {sheet}")

    # Значения записаны.
    for c in VALUE_CELLS:
        v = ws[c].value
        typ = type(v).__name__
        print(f"    {c:>4} = {v!r:<24} (тип {typ})")
        if v is None or (isinstance(v, str) and v.strip() == ""):
            problems.append(f"{sheet}!{c}: пусто")

    # Формулы не перезаписаны.
    for c in FORMULA_CELLS:
        v = ws[c].value
        ok = _is_formula(v)
        print(f"    {c:>4} = {v!r:<40} формула={ok}")
        if not ok:
            problems.append(f"{sheet}!{c}: НЕ формула (перезаписана): {v!r}")

    # Числовой тип у ключевых числовых ячеек (если не прочерк).
    for c in ["D8", "F8", "D14", "F14", "D19", "F19", "D23"]:
        v = ws[c].value
        if v != "—" and not isinstance(v, (int, float)):
            problems.append(f"{sheet}!{c}: не числовой тип и не прочерк: {v!r}")

    return problems


def main() -> int:
    print("=" * 64)
    print(f"ExcelReporter smoke-тест — дата {TEST_DAY.isoformat()}")
    print("=" * 64)

    try:
        service = MetricsService()
        reporter = ExcelReporter()
        saved: dict[str, str] = {}
        computed = {}
        for rn in REPORTS:
            m = service.daily(rn, TEST_DAY)
            computed[rn] = m
            out = os.path.join(OUTPUT_DIR, f"{_safe_name(rn)}_{TEST_DAY.isoformat()}.xlsx")
            reporter.fill_daily(rn, TEST_DAY, m, out)
            saved[rn] = str(out)
            print(f"\nЗаполнен и сохранён: {rn} -> {out}")
    except ODataUnavailableError as exc:
        print("OData НЕДОСТУПЕН (порт 8888 / сеть) — не вина кода:")
        print(f"  {exc}")
        return 2

    all_problems: list[str] = []
    for rn in REPORTS:
        problems = verify_file(saved[rn], SHEET_BY_REPORT[rn])
        all_problems.extend(problems)

    print("\n" + "=" * 64)
    print("СОХРАНЁННЫЕ ФАЙЛЫ:")
    for rn, p in saved.items():
        print(f"  {rn}: {p}")

    print("\n" + "=" * 64)
    if all_problems:
        print("ПРОВАЛЕНО:")
        for p in all_problems:
            print("  -", p)
        return 1
    print("ВСЕ ПРОВЕРКИ ПРОЙДЕНЫ: значения записаны, формулы сохранены.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
