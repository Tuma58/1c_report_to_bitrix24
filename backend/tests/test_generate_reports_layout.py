"""Regression tests for report workbook layout and outgoing titles."""
from __future__ import annotations

import os
import sys
from datetime import date, timedelta

from openpyxl import Workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

from generate_reports import DAILY_SHEETS, _prune_workbook_for_mode, _report_message  # noqa: E402


class FakeService:
    @staticmethod
    def _week_bounds(any_date: date) -> tuple[date, date]:
        start = any_date - timedelta(days=any_date.weekday())
        return start, start + timedelta(days=7)


def main() -> int:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name in [
        "СПРАВОЧНИК РУКОВОДИТЕЛЕЙ",
        "Арсенал (Д)",
        "Арсенал (Н)",
        "Реф.сервис (Д)",
        "Реф.сервис (Н)",
        "Шаркер (Д)",
        "Шаркер (Н)",
        "ЦКР (Д)",
        "ЦКР (Н)",
        "Мойка (Д)",
        "Мойка ул. Строителей (Н)",
        "Мойка Ульяновская (Н)",
    ]:
        wb.create_sheet(sheet_name)

    _prune_workbook_for_mode(wb, "daily")
    assert set(wb.sheetnames) == DAILY_SHEETS, wb.sheetnames

    service = FakeService()
    assert _report_message("daily", service, date(2026, 7, 22), date(2026, 7, 20)) == (
        "Ежедневный отчёт АТЦ за 22.07.2026"
    )
    assert _report_message("weekly", service, date(2026, 7, 22), date(2026, 7, 22)) == (
        "Еженедельный отчёт АТЦ за период 20.07.2026–26.07.2026"
    )

    print("generate reports layout ok")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
