"""Тест Инкремента 7: дневные формы длинных ремонтов ШАРКЕР и ЦКР.

За 2026-07-02 считает ShopMetrics для «Шаркер» и «ЦКР», заполняет листы
«Шаркер (Д)» и «ЦКР (Д)» и сохраняет в output/. Затем ОТКРЫВАЕТ файлы и
проверяет:
- пайплайн Шаркера: active ≈ 40, overdue ≈ 114 (сверка с разведкой);
- payments_period возвращает число (может быть 0) без ошибок; статья резолвится;
- нормочасы/выработка ФАКТ берутся из дневных показателей 1С, если они есть;
- BLOCKED-ячейки (стоимость выработки, страховые ЗН) записаны прочерком «—»;
- формулы колонки H целы.

Только GET из 1С; запись только в xlsx. Если порт недоступен — код 2.
"""
from __future__ import annotations

import os
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

from openpyxl import load_workbook  # noqa: E402

from odata_client import ODataUnavailableError  # noqa: E402
from metrics import MetricsService  # noqa: E402
from excel_reporter import ExcelReporter  # noqa: E402

DAY = date(2026, 7, 2)
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "output"
DASH = "—"


def _fmt(v):
    if v is None:
        return "—"
    if isinstance(v, (int, float)):
        return f"{v:,.2f}".replace(",", " ")
    return str(v)


def _is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


def _is_dash(v) -> bool:
    return v == DASH


def _is_value(v) -> bool:
    return isinstance(v, (int, float))


# Карта FACT/BLOCKED ячеек и формул H по листам.
SHOP = {
    "Шаркер": {
        "sheet": "Шаркер (Д)",
        "fact_cells": {"normhours": "F8", "output_per_master": "F9"},
        "blocked_dash": ["F24", "F25"],   # страховые ФАКТ
        "pipeline": {"active": "D13", "overdue": "D16"},
        "payments_fact": "F23",
        "h_formulas": ["H8", "H9", "H21", "H22", "H23", "H24", "H25"],
    },
    "ЦКР": {
        "sheet": "ЦКР (Д)",
        "fact_cells": {"normhours": "F8", "output_per_master": "F9"},
        "blocked_dash": ["D10", "F10", "F25", "F26"],  # стоимость выработки + страховые
        "pipeline": {"active": "D14", "overdue": "D17"},
        "payments_fact": "F24",
        "h_formulas": ["H8", "H9", "H10", "H22", "H23", "H24", "H25", "H26"],
    },
}


def main() -> int:
    try:
        service = MetricsService()
    except ODataUnavailableError as exc:
        print("OData НЕДОСТУПЕН (порт / сеть) — не вина кода:")
        print(f"  {exc}")
        return 2

    reporter = ExcelReporter()
    failures: list[str] = []
    passed: list[str] = []

    # Проверка резолва статьи ДДС.
    art = service.references.dds_article("Оплата от покупателя")
    (passed if art else failures).append(f"статья 'Оплата от покупателя' резолвится ({art})")

    results = {}
    for name, cfg in SHOP.items():
        print("=" * 70)
        print(f"{name} — ДНЕВНОЙ ОТЧЁТ за {DAY.isoformat()}")
        print("=" * 70)

        m = service.daily_shop(name, DAY)
        results[name] = m

        print("  ВЫРАБОТКА:")
        print(f"    Нормочасы   ПЛАН={_fmt(m.normhours.plan)}  ФАКТ={_fmt(m.normhours.fact)}")
        print(f"    Выработка   ПЛАН={_fmt(m.output_per_master.plan)}  ФАКТ={_fmt(m.output_per_master.fact)}")
        if name == "ЦКР":
            print(f"    Стоимость выработки ПЛАН={_fmt(m.output_cost.plan)}  ФАКТ={_fmt(m.output_cost.fact)} (BLOCKED)")
        print("  ПАЙПЛАЙН (значения 1С):")
        print(f"    Активных ЗН          = {m.active}")
        print(f"    Готовы к выдаче сег. = {m.ready_today}")
        print(f"    Ожидают ЗЧ           = {m.awaiting_parts}")
        print(f"    Просрочены           = {m.overdue}")
        print(f"    План. закрыть неделя = {m.planned_close_week}")
        print("  ФИНАНСЫ:")
        print(f"    Закрыто ЗН  ПЛАН={_fmt(m.closed_orders.plan)}  ФАКТ={_fmt(m.closed_orders.fact)}")
        print(f"    Выручка     ПЛАН={_fmt(m.revenue_closed.plan)}  ФАКТ={_fmt(m.revenue_closed.fact)}")
        print(f"    Оплаты      ФАКТ={_fmt(m.payments.fact)}")
        print(f"    Страховые   count={_fmt(m.insurance_count.fact)}  sum={_fmt(m.insurance_sum.fact)} (BLOCKED)")

        # payments — число без ошибок.
        (passed if isinstance(m.payments.fact, (int, float)) else failures).append(
            f"{name}: payments_period число = {_fmt(m.payments.fact)}"
        )

        # Заполняем и сохраняем.
        out = OUTPUT_DIR / f"shop_{name}_{DAY.isoformat()}.xlsx"
        saved = reporter.fill_daily_shop(name, m, out)
        print(f"  Сохранён: {saved}")

        wb = load_workbook(saved, data_only=False)
        ws = wb[cfg["sheet"]]
        print(f"  [файл] C3={ws['C3'].value!r}  C4={ws['C4'].value!r}")

        # Пайплайн в колонке D записан значениями.
        for key, cc in cfg["pipeline"].items():
            v = ws[cc].value
            expect = getattr(m, key)
            ok = _is_value(v) and abs(v - expect) < 0.01
            (passed if ok else failures).append(
                f"{name}: {key} {cc}={_fmt(v)} (метрика {expect})"
            )

        # ФАКТ-ячейки либо число из 1С, либо «—», если дневной факт не введён.
        for attr, cc in cfg["fact_cells"].items():
            metric = getattr(m, attr)
            v = ws[cc].value
            ok = (_is_value(v) and metric.fact is not None and abs(v - metric.fact) < 0.01) or (
                _is_dash(v) and metric.fact is None
            )
            (passed if ok else failures).append(
                f"{name}: {attr} {cc}={_fmt(v)} (метрика {_fmt(metric.fact)})"
            )

        # BLOCKED-ячейки = «—».
        for cc in cfg["blocked_dash"]:
            v = ws[cc].value
            (passed if _is_dash(v) else failures).append(
                f"{name}: BLOCKED {cc}=«—» ({v!r})"
            )

        # payments ФАКТ записан в F (число).
        pv = ws[cfg["payments_fact"]].value
        (passed if _is_value(pv) else failures).append(
            f"{name}: оплаты {cfg['payments_fact']}={_fmt(pv)} (число)"
        )

        # Формулы H целы.
        for cc in cfg["h_formulas"]:
            v = ws[cc].value
            (passed if _is_formula(v) else failures).append(
                f"{name}: {cc} — формула цела ({v!r})"
            )

    # Сверка с разведкой (Шаркер).
    sh = results["Шаркер"]
    (passed if abs(sh.active - 40) <= 3 else failures).append(
        f"Шаркер active≈40: получено {sh.active}"
    )
    (passed if abs(sh.overdue - 114) <= 6 else failures).append(
        f"Шаркер overdue≈114: получено {sh.overdue}"
    )

    # ------------------------------------------------------------- ОТЧЁТ
    print("=" * 70)
    print("РЕЗУЛЬТАТЫ ПРОВЕРОК")
    print("=" * 70)
    print(f"ПРОШЛО ({len(passed)}):")
    for p in passed:
        print(f"  [OK] {p}")
    print(f"НЕ ПРОШЛО ({len(failures)}):")
    for f in failures:
        print(f"  [FAIL] {f}")

    print("-" * 70)
    print("ИТОГ Шаркер за 02.07.2026:")
    print(f"  active={sh.active}  ready_today={sh.ready_today}  awaiting_parts={sh.awaiting_parts}"
          f"  overdue={sh.overdue}  planned_close_week={sh.planned_close_week}")
    print(f"  оплаты={_fmt(sh.payments.fact)}  выручка={_fmt(sh.revenue_closed.fact)}"
          f"  закрыто={_fmt(sh.closed_orders.fact)}")

    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
