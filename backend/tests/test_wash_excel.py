"""Тест Инкремента 6: заполнение форм МОЕК (дневная + недельная).

За 2026-07-01 (день) и неделю июля считает wash-метрики обеих моек
(«Мойка на ул. Строителей» → «Автомойка»; «Мойка на ул. Ульяновская» →
«Мойка Ульяновская»), заполняет лист «Мойка (Д)» и два недельных листа
отдельно по мойкам, сохраняет
в output/. Затем ОТКРЫВАЕТ файлы и проверяет:
- дневной: оба блока заполнены (D8/F8,D9/F9; D14/F14,D15/F15);
- недельный: раздел 1 отдельно по мойке (D9/F9,D10/F10), раздел 2 по дням (D15:F21);
- SUM(D15:D21)=машин недели, SUM(F15:F21)=выручка недели;
- формульные ячейки (D10,D16 дневного; D11,H15,D22 недельного) остались формулами.

Только GET из 1С; запись только в xlsx. Если порт недоступен — код 2.
"""
from __future__ import annotations

import os
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

from openpyxl import load_workbook  # noqa: E402

from odata_client import ODataUnavailableError  # noqa: E402
from metrics import MetricsService  # noqa: E402
from excel_reporter import ExcelReporter  # noqa: E402


DAY = date(2026, 7, 1)          # дневной отчёт
WEEK_DAY = date(2026, 7, 1)     # опорная дата недели
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "output"

STROIT = "Мойка на ул. Строителей"   # -> Автомойка
ULYAN = "Мойка на ул. Ульяновская"   # -> Мойка Ульяновская
NAMES = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]


def _fmt(v):
    if v is None:
        return "—"
    if isinstance(v, (int, float)):
        return f"{v:,.2f}".replace(",", " ")
    return str(v)


def _is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


def _is_value(v) -> bool:
    return isinstance(v, (int, float))


def main() -> int:
    try:
        service = MetricsService()
    except ODataUnavailableError as exc:
        print("OData НЕДОСТУПЕН (порт / сеть) — не вина кода:")
        print(f"  {exc}")
        return 2

    reporter = ExcelReporter()
    failures: list[str] = []
    passed: list[str] = []

    # ----------------------------------------------------------- ДНЕВНОЙ
    print("=" * 70)
    print(f"МОЙКИ — ДНЕВНОЙ ОТЧЁТ за {DAY.isoformat()}")
    print("=" * 70)

    d_stroit = service.daily_wash(STROIT, DAY)
    d_ulyan = service.daily_wash(ULYAN, DAY)

    for nm, m in ((STROIT, d_stroit), (ULYAN, d_ulyan)):
        print(f"  {nm}")
        print(f"    Машин   ПЛАН={_fmt(m.cars.plan)}  ФАКТ={_fmt(m.cars.fact)}")
        print(f"    Выручка ПЛАН={_fmt(m.revenue.plan)}  ФАКТ={_fmt(m.revenue.fact)}")

    daily_out = OUTPUT_DIR / f"wash_daily_{DAY.isoformat()}.xlsx"
    saved_d = reporter.fill_daily_wash(DAY, d_stroit, d_ulyan, daily_out)
    print(f"  Сохранён: {saved_d}")

    wb = load_workbook(saved_d, data_only=False)
    ws = wb["Мойка (Д)"]
    print(f"  [файл] C3={ws['C3'].value!r}  C4={ws['C4'].value!r}")

    # Блоки заполнены (значения, не формулы).
    for cc, expect in [
        ("F8", d_stroit.cars.fact), ("F9", d_stroit.revenue.fact),
        ("F14", d_ulyan.cars.fact), ("F15", d_ulyan.revenue.fact),
    ]:
        v = ws[cc].value
        ok = _is_value(v) and abs(v - expect) < 0.01
        (passed if ok else failures).append(f"дневной {cc}={_fmt(v)} (ожид {_fmt(expect)})")

    # Формулы Ср.чек целы.
    for cc in ["D10", "F10", "D16", "F16"]:
        v = ws[cc].value
        (passed if _is_formula(v) else failures).append(
            f"дневной {cc} — формула цела ({v!r})"
        )

    # ----------------------------------------------------------- НЕДЕЛЬНЫЙ
    print("=" * 70)
    print(f"МОЙКИ — НЕДЕЛЬНЫЙ ОТЧЁТ (опорная {WEEK_DAY.isoformat()})")
    print("=" * 70)

    weekly_cases = [
        (STROIT, "Мойка ул. Строителей (Н)", service.weekly_wash(STROIT, WEEK_DAY),
         service.week_wash_daily_breakdown(STROIT, WEEK_DAY)),
        (ULYAN, "Мойка Ульяновская (Н)", service.weekly_wash(ULYAN, WEEK_DAY),
         service.week_wash_daily_breakdown(ULYAN, WEEK_DAY)),
    ]
    saved_weekly = []
    for report_name, sheet_name, weekly, breakdown in weekly_cases:
        assert len(breakdown) == 7, "разбивка должна быть 7 дней"
        print(f"  {report_name}: {weekly.period_start.isoformat()} .. {weekly.period_end.isoformat()}")
        print(f"    Машин  ПЛАН={_fmt(weekly.cars.plan)}  ФАКТ={_fmt(weekly.cars.fact)}")
        print(f"    Выручка ПЛАН={_fmt(weekly.revenue.plan)}  ФАКТ={_fmt(weekly.revenue.fact)}")
        for nm, dm in zip(NAMES, breakdown):
            print(f"    {nm} {dm.period_start.isoformat()}: машин={_fmt(dm.cars.fact)} "
                  f"выручка={_fmt(dm.revenue.fact)} операторов={dm.executors_count}")

        sum_cars_days = sum((dm.cars.fact or 0.0) for dm in breakdown)
        sum_rev_days = sum((dm.revenue.fact or 0.0) for dm in breakdown)

        weekly_out = OUTPUT_DIR / f"wash_weekly_{sheet_name}_{WEEK_DAY.isoformat()}.xlsx"
        saved_w = reporter.fill_weekly_wash(report_name, weekly, breakdown, weekly_out)
        saved_weekly.append(saved_w)
        print(f"  Сохранён: {saved_w}")

        wb2 = load_workbook(saved_w, data_only=False)
        ws2 = wb2[sheet_name]
        print(f"  [файл] {sheet_name}: C3={ws2['C3'].value!r}  C4={ws2['C4'].value!r}")

        for cc, expect in [("F9", weekly.cars.fact), ("F10", weekly.revenue.fact)]:
            v = ws2[cc].value
            ok = _is_value(v) and abs(v - expect) < 0.01
            (passed if ok else failures).append(
                f"{sheet_name} {cc}={_fmt(v)} (ожид {_fmt(expect)})"
            )

        d_cells = sum((ws2[f"D{15+i}"].value or 0.0) for i in range(7)
                      if _is_value(ws2[f"D{15+i}"].value))
        f_cells = sum((ws2[f"F{15+i}"].value or 0.0) for i in range(7)
                      if _is_value(ws2[f"F{15+i}"].value))

        ok_d = abs(d_cells - weekly.cars.fact) < 0.01 and abs(d_cells - sum_cars_days) < 0.01
        ok_f = abs(f_cells - weekly.revenue.fact) < 0.01 and abs(f_cells - sum_rev_days) < 0.01
        (passed if ok_d else failures).append(
            f"{sheet_name} SUM(D15:D21)={_fmt(d_cells)} = машин недели {_fmt(weekly.cars.fact)}"
        )
        (passed if ok_f else failures).append(
            f"{sheet_name} SUM(F15:F21)={_fmt(f_cells)} = выручка недели {_fmt(weekly.revenue.fact)}"
        )

        for cc in ["D11", "F11", "H15", "H21", "D22", "F22", "H22"]:
            v = ws2[cc].value
            (passed if _is_formula(v) else failures).append(
                f"{sheet_name} {cc} — формула цела ({v!r})"
            )

    # ----------------------------------------------------------- ОТЧЁТ
    print("=" * 70)
    print("РЕЗУЛЬТАТЫ ПРОВЕРОК")
    print("=" * 70)
    print(f"ПРОШЛО ({len(passed)}):")
    for p in passed:
        print(f"  [OK] {p}")
    print(f"НЕ ПРОШЛО ({len(failures)}):")
    for f in failures:
        print(f"  [FAIL] {f}")

    print("-" * 70)
    print("ФАКТ за 01.07.2026:")
    print(f"  {STROIT}: машин={_fmt(d_stroit.cars.fact)}  выручка={_fmt(d_stroit.revenue.fact)}")
    print(f"  {ULYAN}: машин={_fmt(d_ulyan.cars.fact)}  выручка={_fmt(d_ulyan.revenue.fact)}")
    print(f"  xlsx дневной:   {saved_d}")
    for saved_w in saved_weekly:
        print(f"  xlsx недельный: {saved_w}")

    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
