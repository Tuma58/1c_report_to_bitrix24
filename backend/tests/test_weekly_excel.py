"""Тест Инкремента 5: заполнение НЕДЕЛЬНЫХ Excel-форм (Арсенал, Реф. Сервис).

За неделю июля 2026 (план есть) считает weekly() + week_daily_breakdown() для
«Арсенал» и «Реф. Сервис», заполняет листы (Н), сохраняет в output/. Затем
ОТКРЫВАЕТ сохранённые файлы и проверяет:
- C3/C4 заполнены;
- D9/F9, D11/F11, D15/F15, D16/F16, D20/F20 — значения;
- раздел 3 (D24:H30) заполнен по дням;
- сумма D24:D30 ≈ weekly.revenue.fact;
- формульные ячейки (D17,D18,D19,I24,D31,H9) остались формулами (начинаются с '=').

Только GET из 1С; запись только в xlsx. Если порт недоступен — код 2.
"""
from __future__ import annotations

import os
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

from openpyxl import load_workbook  # noqa: E402

from odata_client import ODataUnavailableError  # noqa: E402
from metrics import MetricsService  # noqa: E402
from excel_reporter import ExcelReporter  # noqa: E402


WEEK_DAY = date(2026, 7, 1)  # среда — ISO-неделя 29.06–05.07.2026
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "output"
REPORTS = ["Арсенал", "Реф. Сервис"]


def _fmt(v):
    if v is None:
        return "—"
    if isinstance(v, (int, float)):
        return f"{v:,.2f}".replace(",", " ")
    return str(v)


def _approx(a, b, tol) -> bool:
    return a is not None and isinstance(a, (int, float)) and abs(a - b) <= tol


def _is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


def _is_value(v) -> bool:
    return isinstance(v, (int, float))


def main() -> int:
    try:
        service = MetricsService()
    except ODataUnavailableError as exc:
        print("OData НЕДОСТУПЕН (порт / сеть) — не вина кода:")
        print(f"  {exc}")
        return 2

    reporter = ExcelReporter()
    failures: list[str] = []
    passed: list[str] = []

    try:
        for report_name in REPORTS:
            print("=" * 70)
            print(f"АТЦ: {report_name} — неделя {WEEK_DAY.isoformat()}")
            print("=" * 70)

            wk = service.weekly(report_name, WEEK_DAY)
            days = service.week_daily_breakdown(report_name, WEEK_DAY)

            assert len(days) == 7, "week_daily_breakdown должен вернуть 7 дней"
            # Проверка порядка: день 0 = week_start (Пн), день 6 = week_end (Вс).
            if days[0].day == wk.week_start and days[6].day == wk.week_end:
                passed.append(f"{report_name}: разбивка 7 дней Пн..Вс в порядке")
            else:
                failures.append(f"{report_name}: неверный порядок дней разбивки")

            print(f"  Неделя: {wk.week_start.isoformat()} .. {wk.week_end.isoformat()}")
            print(f"  Нормочасы   ПЛАН={_fmt(wk.normhours.plan)}  ФАКТ={_fmt(wk.normhours.fact)}")
            print(f"  Закрыто ЗН  ПЛАН={_fmt(wk.closed_orders.plan)}  ФАКТ={_fmt(wk.closed_orders.fact)}")
            print(f"  Выручка     ПЛАН={_fmt(wk.revenue.plan)}  ФАКТ={_fmt(wk.revenue.fact)}")
            print(f"  Себест.ЗЧ   ПЛАН={_fmt(wk.cost.plan)}  ФАКТ={_fmt(wk.cost.fact)}")
            print(f"  Наценка %   ПЛАН={_fmt(wk.markup_pct.plan)}  ФАКТ={_fmt(wk.markup_pct.fact)}")
            print("  Разбивка по дням (ФАКТ):")
            names = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
            for nm, dm in zip(names, days):
                print(f"    {nm} {dm.day.isoformat()}: выручка={_fmt(dm.revenue.fact)} "
                      f"нормочасы={_fmt(dm.normhours.fact)} ЗН={_fmt(dm.closed_orders.fact)}")

            out_path = OUTPUT_DIR / f"weekly_{report_name}_{WEEK_DAY.isoformat()}.xlsx"
            saved = reporter.fill_weekly(report_name, wk, days, out_path)
            print(f"  Сохранён: {saved}")

            # --- Открываем и проверяем сохранённый файл ---
            wb = load_workbook(saved, data_only=False)
            sheet = "Арсенал (Н)" if report_name == "Арсенал" else "Реф.сервис (Н)"
            ws = wb[sheet]

            c3, c4 = ws["C3"].value, ws["C4"].value
            print(f"  [файл] C3={c3!r}  C4={c4!r}")

            (passed if c3 == report_name else failures).append(f"{report_name}: C3 заполнен")
            (passed if isinstance(c4, str) and "–" in c4 else failures).append(
                f"{report_name}: C4 неделя-строка"
            )

            # Разделы 1–2: значения.
            value_cells = ["D9", "F9", "D11", "F11", "D15", "F15", "D16", "F16", "D20", "F20"]
            for cc in value_cells:
                v = ws[cc].value
                ok = _is_value(v) or v == "—"
                # для наценки/плана возможен «—»; главное — не формула и не пусто
                ok = ok and not _is_formula(v)
                (passed if ok else failures).append(f"{report_name}: {cc} значение ({_fmt(v)})")

            # Раздел 3: D24:H30 заполнены (значение или «—»).
            sec3_ok = True
            total_d = 0.0
            for row in range(24, 31):
                for col in ("D", "F", "H"):
                    v = ws[f"{col}{row}"].value
                    if _is_formula(v) or v is None:
                        sec3_ok = False
                for col in ("D",):
                    v = ws[f"{col}{row}"].value
                    if _is_value(v):
                        total_d += v
            (passed if sec3_ok else failures).append(f"{report_name}: раздел 3 D24:H30 заполнен")

            # Сумма D24:D30 ≈ weekly.revenue.fact.
            (passed if _approx(total_d, wk.revenue.fact, max(1.0, wk.revenue.fact * 0.001)) else failures).append(
                f"{report_name}: SUM(D24:D30)={_fmt(total_d)} ≈ revenue.fact={_fmt(wk.revenue.fact)}"
            )

            # Формульные ячейки целы.
            formula_cells = ["D17", "F17", "D18", "F18", "D19", "F19",
                             "I24", "I30", "D31", "F31", "H31", "H9", "H15"]
            intact = []
            broken = []
            for fc in formula_cells:
                v = ws[fc].value
                (intact if _is_formula(v) else broken).append(f"{fc}={v!r}")
            if not broken:
                passed.append(f"{report_name}: формульные ячейки целы ({len(intact)} шт.)")
            else:
                failures.append(f"{report_name}: формулы затёрты: {broken}")
            print(f"  [файл] формулы целы: {', '.join(intact)}")
            print()

    except ODataUnavailableError as exc:
        print("OData стал недоступен во время теста — не вина кода:")
        print(f"  {exc}")
        return 2

    print("=" * 70)
    print("ИТОГ КРИТЕРИЕВ")
    print("=" * 70)
    for p in passed:
        print(f"  [PASS] {p}")
    for f in failures:
        print(f"  [FAIL] {f}")

    if failures:
        print(f"\nНЕ ПРОШЛО: {len(failures)} из {len(passed) + len(failures)}")
        return 1
    print(f"\nВСЕ КРИТЕРИИ ПРОШЛИ: {len(passed)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
